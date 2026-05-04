"""‌⁠‍Build a multi-sheet Excel that joins demo-form signups (marketing
   funnel) with actual ERP user accounts and their activity counts.

   Sheets:
     1. Signups + Activity — every form signup, joined to the user row
        in the ERP DB by email. Tells you who actually activated and
        used the product.
     2. Active Users — every user in the ERP DB (incl. self-registered
        outside the marketing form), with project/BOQ/chat counts.
     3. Funnel summary — counts at each step of the funnel.

   Activity columns:
     - Logged in?     — yes if last_login_at is not NULL
     - Last login     — timestamp of latest login
     - Projects       — count of oe_projects_project rows owned
     - BOQ actions    — count of oe_boq_activity_log entries by user
     - Chat sessions  — count of oe_erp_chat_session rows by user
     - Audit actions  — count of oe_core_audit_log entries by user
     - Total activity — projects + boq + chat + audit
"""
import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DESKTOP = Path('C:/Users/Artem Boiko/Desktop')
JSONL = DESKTOP / 'demo-registrations.jsonl'
TSV = DESKTOP / 'users_activity.tsv'
DST = DESKTOP / 'demo-registrations.xlsx'

# ─── 1. Load demo-form signups (tolerant parser) ────────────────────────
signups = []
with JSONL.open(encoding='utf-8') as f:
    decoder = json.JSONDecoder()
    for line in f:
        line = line.strip()
        if not line:
            continue
        idx = 0
        while idx < len(line):
            try:
                obj, end = decoder.raw_decode(line, idx)
            except json.JSONDecodeError:
                break
            signups.append(obj)
            idx = end
            while idx < len(line) and line[idx] in ' \t,':
                idx += 1

# ─── 2. Load ERP user activity (TSV pulled via SSH) ─────────────────────
users = []
with TSV.open(encoding='utf-8') as f:
    header = f.readline().rstrip('\n').split('|')
    for line in f:
        cells = line.rstrip('\n').split('|')
        if len(cells) != len(header):
            continue
        users.append(dict(zip(header, cells)))

users_by_email = {u['email'].strip().lower(): u for u in users}

def _user_activity(email: str) -> dict:
    u = users_by_email.get(email.strip().lower())
    if not u:
        return {
            'activated': '',
            'logged_in': '',
            'last_login': '',
            'projects': '',
            'boq_actions': '',
            'chat_sessions': '',
            'audit_actions': '',
            'total_activity': '',
        }
    p = int(u['projects_created'] or 0)
    b = int(u['boq_actions'] or 0)
    c = int(u['chat_sessions'] or 0)
    a = int(u['audit_actions'] or 0)
    return {
        'activated': 'yes',
        'logged_in': 'yes' if u['last_login_at'] else 'no',
        'last_login': u['last_login_at'] or '',
        'projects': p,
        'boq_actions': b,
        'chat_sessions': c,
        'audit_actions': a,
        'total_activity': p + b + c + a,
    }

# ─── 3. Workbook setup helpers ──────────────────────────────────────────
wb = Workbook()

HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='1F2937')
SUBTLE_FILL = PatternFill('solid', fgColor='F3F4F6')

def _style_header(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical='center', horizontal='left')

def _autosize(ws, headers, cap_per_col=None):
    cap_per_col = cap_per_col or {}
    for col_idx, hdr in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        max_len = len(str(hdr))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            v = row[0].value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        cap = cap_per_col.get(hdr, 50)
        ws.column_dimensions[letter].width = min(max_len + 2, cap)

# ─── 4. Sheet 1: Signups + activity ─────────────────────────────────────
ws1 = wb.active
ws1.title = 'Signups + Activity'

headers1 = [
    '#', 'Server time (UTC)', 'First name', 'Last name', 'Email',
    'Company', 'Role (form)', 'Company size', 'Language', 'IP',
    'Activated?', 'Logged in?', 'Last login (UTC)',
    'Projects', 'BOQ actions', 'Chat sessions', 'Audit actions',
    'Total activity',
]
ws1.append(headers1)
_style_header(ws1, len(headers1))

# Newest first
signups.sort(key=lambda r: (r.get('server_time') or r.get('timestamp') or ''), reverse=True)

for i, r in enumerate(signups, 1):
    email = (r.get('email') or '').strip()
    act = _user_activity(email)
    ws1.append([
        i,
        r.get('server_time', ''),
        r.get('firstName', ''),
        r.get('lastName', ''),
        email,
        r.get('company', ''),
        r.get('role', ''),
        r.get('companySize', ''),
        r.get('language', ''),
        r.get('ip', ''),
        act['activated'],
        act['logged_in'],
        act['last_login'],
        act['projects'],
        act['boq_actions'],
        act['chat_sessions'],
        act['audit_actions'],
        act['total_activity'],
    ])

ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = ws1.dimensions
_autosize(ws1, headers1)

# ─── 5. Sheet 2: Active Users (DB-side) ─────────────────────────────────
ws2 = wb.create_sheet('Active Users')
headers2 = [
    '#', 'Email', 'Full name', 'Role', 'Locale',
    'Created (UTC)', 'Last login (UTC)',
    'Projects', 'BOQ actions', 'Chat sessions', 'Audit actions',
    'Total activity',
]
ws2.append(headers2)
_style_header(ws2, len(headers2))

users_sorted = sorted(
    users,
    key=lambda u: (
        int(u['projects_created'] or 0)
        + int(u['boq_actions'] or 0)
        + int(u['chat_sessions'] or 0)
        + int(u['audit_actions'] or 0)
    ),
    reverse=True,
)

for i, u in enumerate(users_sorted, 1):
    p = int(u['projects_created'] or 0)
    b = int(u['boq_actions'] or 0)
    c = int(u['chat_sessions'] or 0)
    a = int(u['audit_actions'] or 0)
    ws2.append([
        i,
        u['email'],
        u['full_name'],
        u['role'],
        u['locale'],
        u['created_at'],
        u['last_login_at'] or '',
        p, b, c, a, p + b + c + a,
    ])

ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = ws2.dimensions
_autosize(ws2, headers2)

# ─── 6. Sheet 3: Funnel summary ─────────────────────────────────────────
ws3 = wb.create_sheet('Funnel')

ws3.append(['Marketing → product funnel'])
ws3.cell(row=1, column=1).font = Font(bold=True, size=14)
ws3.append([])

total_signups = len(signups)
distinct_signup_emails = {(s.get('email') or '').strip().lower() for s in signups if s.get('email')}
distinct_signup_emails.discard('')

activated_emails = distinct_signup_emails & set(users_by_email)
logged_in_emails = {e for e in activated_emails if users_by_email[e]['last_login_at']}
used_product_emails = {
    e for e in activated_emails
    if (int(users_by_email[e]['projects_created'] or 0)
        + int(users_by_email[e]['boq_actions'] or 0)
        + int(users_by_email[e]['chat_sessions'] or 0)
        + int(users_by_email[e]['audit_actions'] or 0)) > 0
}

rows = [
    ('Total form signups (raw rows)', total_signups),
    ('Distinct emails on the form', len(distinct_signup_emails)),
    ('Activated (also exist in ERP DB)', len(activated_emails)),
    ('Logged in at least once', len(logged_in_emails)),
    ('Did anything in the product (>0 actions)', len(used_product_emails)),
    ('', ''),
    ('Total ERP users (incl. self-registered)', len(users)),
    ('ERP users with last_login_at set', sum(1 for u in users if u['last_login_at'])),
    ('ERP users who created at least one project', sum(1 for u in users if int(u['projects_created'] or 0) > 0)),
    ('Total projects across all users', sum(int(u['projects_created'] or 0) for u in users)),
    ('Total BOQ actions across all users', sum(int(u['boq_actions'] or 0) for u in users)),
    ('Total chat sessions across all users', sum(int(u['chat_sessions'] or 0) for u in users)),
]

ws3.append(['Metric', 'Value'])
_style_header(ws3, 2, row=ws3.max_row)
for label, val in rows:
    ws3.append([label, val])

ws3.column_dimensions['A'].width = 48
ws3.column_dimensions['B'].width = 14

wb.save(DST)
print(f'wrote {DST}')
print(f'signups: {total_signups}, db users: {len(users)}, activated: {len(activated_emails)}')
