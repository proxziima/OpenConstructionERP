"""‌⁠‍Convert demo-registrations.jsonl into a polished Excel workbook
   with sortable headers, frozen top row, autofit columns, and a
   summary sheet (counts by language / role / company size)."""
import json
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = Path('C:/Users/Artem Boiko/Desktop/demo-registrations.jsonl')
DST = Path('C:/Users/Artem Boiko/Desktop/demo-registrations.xlsx')

rows = []
with SRC.open(encoding='utf-8') as f:
    for line_no, line in enumerate(f, 1):
        line = line.strip()
        if not line:
            continue
        # Some lines contain two concatenated records (write race on the
        # server). Use raw_decode in a loop to recover both.
        decoder = json.JSONDecoder()
        idx = 0
        while idx < len(line):
            try:
                obj, end = decoder.raw_decode(line, idx)
            except json.JSONDecodeError as exc:
                print(f'!! line {line_no} skipped at offset {idx}: {exc}')
                break
            rows.append(obj)
            idx = end
            while idx < len(line) and line[idx] in ' \t,':
                idx += 1

# Order rows newest first by server_time (fallback to timestamp).
def _sort_key(r):
    return r.get('server_time') or r.get('timestamp') or ''
rows.sort(key=_sort_key, reverse=True)

wb = Workbook()

# ─── Sheet 1: All registrations ─────────────────────────────────────────
ws = wb.active
ws.title = 'Registrations'

headers = [
    '#', 'Server time (UTC)', 'First name', 'Last name', 'Email',
    'Company', 'Role', 'Company size', 'Language', 'IP',
    'Client time', 'User-Agent',
]
ws.append(headers)

# Header style
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1F2937')
for col_idx, _ in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(vertical='center', horizontal='left')

for i, r in enumerate(rows, 1):
    ws.append([
        i,
        r.get('server_time', ''),
        r.get('firstName', ''),
        r.get('lastName', ''),
        r.get('email', ''),
        r.get('company', ''),
        r.get('role', ''),
        r.get('companySize', ''),
        r.get('language', ''),
        r.get('ip', ''),
        r.get('timestamp', ''),
        r.get('userAgent', ''),
    ])

ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

# Auto-fit-ish: clamp width per column based on max cell length.
for col_idx in range(1, len(headers) + 1):
    letter = get_column_letter(col_idx)
    max_len = len(str(headers[col_idx - 1]))
    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
        v = row[0].value
        if v is None:
            continue
        max_len = max(max_len, len(str(v)))
    # Cap UA column to keep workbook readable.
    cap = 60 if headers[col_idx - 1] == 'User-Agent' else 50
    ws.column_dimensions[letter].width = min(max_len + 2, cap)

# ─── Sheet 2: Summary ───────────────────────────────────────────────────
ws2 = wb.create_sheet('Summary')

def _emit(title, counter, sheet):
    sheet.append([title])
    sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True, size=13)
    sheet.append(['Value', 'Count'])
    for c in (sheet.cell(row=sheet.max_row, column=i) for i in (1, 2)):
        c.font = header_font
        c.fill = header_fill
    for k, v in counter.most_common():
        sheet.append([k or '(empty)', v])
    sheet.append([])

_emit(f'Total signups: {len(rows)}', Counter(), ws2)
ws2.append([])

_emit('By language', Counter(r.get('language', '') for r in rows), ws2)
_emit('By role', Counter(r.get('role', '') for r in rows), ws2)
_emit('By company size', Counter(r.get('companySize', '') for r in rows), ws2)

# Top 20 email domains
domains = Counter()
for r in rows:
    em = (r.get('email') or '').lower()
    if '@' in em:
        domains[em.rsplit('@', 1)[1]] += 1
ws2.append(['Top 20 email domains'])
ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=13)
ws2.append(['Domain', 'Count'])
for c in (ws2.cell(row=ws2.max_row, column=i) for i in (1, 2)):
    c.font = header_font
    c.fill = header_fill
for d, n in domains.most_common(20):
    ws2.append([d, n])

ws2.column_dimensions['A'].width = 32
ws2.column_dimensions['B'].width = 12

wb.save(DST)
print(f'wrote {DST}')
print(f'rows: {len(rows)}')
