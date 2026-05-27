"""Unit tests for the COBie parser."""

import tempfile
from pathlib import Path

import pytest

from app.modules.bim_requirements.parsers.cobie_parser import COBieParser


@pytest.fixture
def parser() -> COBieParser:
    return COBieParser()


def _create_cobie_xlsx(attributes: list[list]) -> Path:
    """Create a minimal COBie Excel file with an Attribute sheet."""
    import openpyxl

    wb = openpyxl.Workbook()
    # Rename default sheet to Attribute
    ws = wb.active
    ws.title = "Attribute"  # type: ignore[union-attr]

    # COBie Attribute headers
    headers = [
        "Name",
        "CreatedBy",
        "CreatedOn",
        "Category",
        "SheetName",
        "RowName",
        "Value",
        "Unit",
        "ExtSystem",
        "ExtObject",
        "ExtIdentifier",
        "Description",
        "AllowedValues",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = h  # type: ignore[union-attr]

    for r_idx, row_data in enumerate(attributes, start=2):
        for c_idx, val in enumerate(row_data, start=1):
            ws.cell(row=r_idx, column=c_idx).value = val  # type: ignore[union-attr]

    # Add a Component sheet to make it look more like COBie
    wb.create_sheet("Component")
    wb.create_sheet("Type")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
    return Path(tmp.name)


class TestCOBieParser:
    """Tests for the COBie parser."""

    def test_basic_attribute(self, parser: COBieParser) -> None:
        """Parse a single COBie attribute row."""
        rows = [
            [
                "FireRating",  # Name
                "user@example.com",  # CreatedBy
                "2024-01-01",  # CreatedOn
                "Fire Safety",  # Category
                "Type",  # SheetName
                "Wall-001",  # RowName
                "REI60",  # Value
                "",  # Unit
                "",  # ExtSystem
                "",  # ExtObject
                "",  # ExtIdentifier
                "Fire resistance",  # Description
                "REI60;REI90",  # AllowedValues
            ]
        ]
        path = _create_cobie_xlsx(rows)
        try:
            result = parser.parse(path)
            assert result.success
            assert len(result.requirements) == 1

            req = result.requirements[0]
            assert req.property_name == "FireRating"
            assert req.element_filter.get("cobie_sheet") == "Type"
            assert req.element_filter.get("cobie_row_name") == "Wall-001"
            assert req.constraint_def.get("value") == "REI60"
            assert req.constraint_def.get("enum") == ["REI60", "REI90"]
            assert req.context.get("category") == "Fire Safety"
            assert req.context.get("description") == "Fire resistance"
        finally:
            path.unlink(missing_ok=True)

    def test_multiple_attributes(self, parser: COBieParser) -> None:
        """Parse multiple COBie attribute rows."""
        rows = [
            ["FireRating", "", "", "", "Type", "Wall-001", "REI60", "", "", "", "", "", ""],
            ["Acoustic", "", "", "", "Type", "Wall-002", "", "dB", "", "", "", "", ""],
            ["Material", "", "", "", "Component", "Beam-001", "Steel", "", "", "", "", "", ""],
        ]
        path = _create_cobie_xlsx(rows)
        try:
            result = parser.parse(path)
            assert result.success
            assert len(result.requirements) == 3
        finally:
            path.unlink(missing_ok=True)

    def test_empty_attribute_sheet(self, parser: COBieParser) -> None:
        """Empty Attribute sheet produces a warning."""
        path = _create_cobie_xlsx([])
        try:
            result = parser.parse(path)
            assert not result.success
            assert len(result.warnings) > 0
        finally:
            path.unlink(missing_ok=True)

    def test_property_group_is_none(self, parser: COBieParser) -> None:
        """COBie does not have property groups -- should be None."""
        rows = [["FireRating", "", "", "", "Type", "Wall-001", "REI60", "", "", "", "", "", ""]]
        path = _create_cobie_xlsx(rows)
        try:
            result = parser.parse(path)
            assert result.success
            assert result.requirements[0].property_group is None
        finally:
            path.unlink(missing_ok=True)

    def test_allowed_values_parsing(self, parser: COBieParser) -> None:
        """AllowedValues with commas are parsed as enum."""
        rows = [["Status", "", "", "", "Type", "Door-001", "", "", "", "", "", "", "Open,Closed,Locked"]]
        path = _create_cobie_xlsx(rows)
        try:
            result = parser.parse(path)
            assert result.success
            assert result.requirements[0].constraint_def.get("enum") == [
                "Open",
                "Closed",
                "Locked",
            ]
        finally:
            path.unlink(missing_ok=True)
