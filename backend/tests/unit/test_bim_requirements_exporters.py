"""Unit tests for the BIM requirements exporters.

Includes round-trip tests: create requirements -> export -> re-import -> verify.
"""

import io

from app.modules.bim_requirements.exporters.excel_exporter import (
    export_excel,
    generate_template,
)
from app.modules.bim_requirements.exporters.ids_exporter import export_ids_xml
from app.modules.bim_requirements.parsers.base import UniversalRequirement
from app.modules.bim_requirements.parsers.ids_parser import IDSParser


def _sample_requirements() -> list[UniversalRequirement]:
    """Create a standard set of test requirements."""
    return [
        UniversalRequirement(
            element_filter={"ifc_class": "IFCWALL"},
            property_group="Pset_WallCommon",
            property_name="FireRating",
            constraint_def={
                "datatype": "IFCLABEL",
                "cardinality": "required",
                "enum": ["REI60", "REI90", "REI120"],
            },
            context={"ifc_version": "IFC4", "use_case": "Fire Safety"},
        ),
        UniversalRequirement(
            element_filter={"ifc_class": "IFCWALL"},
            property_group="Pset_WallCommon",
            property_name="ThermalTransmittance",
            constraint_def={
                "datatype": "IFCREAL",
                "cardinality": "required",
                "min": 0.1,
                "max": 1.0,
            },
            context={"ifc_version": "IFC4", "use_case": "Fire Safety"},
        ),
        UniversalRequirement(
            element_filter={"ifc_class": "IFCDOOR"},
            property_group=None,
            property_name="Description",
            constraint_def={"cardinality": "required", "value": "Standard Door"},
            context={"ifc_version": "IFC4", "use_case": "Door Reqs"},
        ),
    ]


class TestExcelExporter:
    """Tests for the Excel exporter."""

    def test_export_produces_valid_xlsx(self) -> None:
        """Export produces a valid Excel workbook."""
        import openpyxl

        reqs = _sample_requirements()
        content = export_excel(reqs)
        assert isinstance(content, bytes)
        assert len(content) > 0

        wb = openpyxl.load_workbook(io.BytesIO(content))
        assert "Requirements" in wb.sheetnames
        assert "Summary" in wb.sheetnames
        assert "Legend" in wb.sheetnames

        ws = wb["Requirements"]
        # Header row + 3 data rows
        assert ws.max_row >= 4
        # Check header content
        assert ws.cell(row=1, column=1).value == "Element / IFC Class"
        # Check data
        assert ws.cell(row=2, column=1).value == "IFCWALL"
        assert ws.cell(row=2, column=4).value == "FireRating"

    def test_export_german_headers(self) -> None:
        """German language headers are used when language=de."""
        import openpyxl

        reqs = _sample_requirements()[:1]
        content = export_excel(reqs, language="de")
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb["Requirements"]
        assert ws.cell(row=1, column=1).value == "Bauteil / IFC-Klasse"

    def test_template_generation(self) -> None:
        """Template generates a valid Excel file with example data."""
        import openpyxl

        content = generate_template()
        assert isinstance(content, bytes)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb["Requirements"]
        # Template has 5 example rows
        assert ws.max_row >= 6

    def test_summary_sheet_counts(self) -> None:
        """Summary sheet shows correct element counts."""
        import openpyxl

        reqs = _sample_requirements()
        content = export_excel(reqs)
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb["Summary"]
        # Check total row exists
        found_total = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == "Total":
                assert row[1] == 3
                found_total = True
        assert found_total


class TestIDSExporter:
    """Tests for the IDS XML exporter."""

    def test_export_produces_valid_xml(self) -> None:
        """Export produces parseable XML."""
        import xml.etree.ElementTree as ET

        reqs = _sample_requirements()
        xml_str = export_ids_xml(reqs, title="Test Export")
        assert isinstance(xml_str, str)
        assert "<?xml version" in xml_str

        # Must be parseable
        root = ET.fromstring(xml_str)
        assert root is not None

    def test_roundtrip_ids_to_universal_to_ids(self) -> None:
        """Parse IDS -> export IDS -> parse again -> data matches."""
        # Create requirements with known structure
        original_reqs = [
            UniversalRequirement(
                element_filter={"ifc_class": "IFCWALL"},
                property_group="Pset_WallCommon",
                property_name="FireRating",
                constraint_def={
                    "datatype": "IFCLABEL",
                    "cardinality": "required",
                    "enum": ["REI60", "REI90"],
                },
                context={"ifc_version": "IFC4", "use_case": "Roundtrip Test"},
            ),
        ]

        # Export to IDS XML
        xml_str = export_ids_xml(original_reqs, title="Roundtrip Test")

        # Parse back
        parser = IDSParser()
        result = parser.parse(xml_str)
        assert result.success
        assert len(result.requirements) == 1

        # Verify the key fields survive the round-trip
        req = result.requirements[0]
        assert req.element_filter.get("ifc_class") == "IFCWALL"
        assert req.property_group == "Pset_WallCommon"
        assert req.property_name == "FireRating"
        assert req.constraint_def.get("cardinality") == "required"
        assert req.constraint_def.get("datatype") == "IFCLABEL"
        assert set(req.constraint_def.get("enum", [])) == {"REI60", "REI90"}

    def test_grouping_by_element_filter(self) -> None:
        """Requirements with same element_filter are grouped into one specification."""
        import xml.etree.ElementTree as ET

        reqs = _sample_requirements()
        xml_str = export_ids_xml(reqs)
        root = ET.fromstring(xml_str)

        # Count specifications -- IFCWALL reqs should be grouped together
        ns = {"ids": "http://standards.buildingsmart.org/IDS"}
        specs = root.findall(".//ids:specification", ns)
        # Two groups: IFCWALL (2 reqs) + IFCDOOR (1 req)
        assert len(specs) == 2

    def test_attribute_vs_property_facets(self) -> None:
        """Requirements without property_group become <attribute> facets."""
        import xml.etree.ElementTree as ET

        reqs = [
            UniversalRequirement(
                element_filter={"ifc_class": "IFCWALL"},
                property_group=None,
                property_name="Name",
                constraint_def={"cardinality": "required"},
                context={"ifc_version": "IFC4", "use_case": "Test"},
            ),
        ]
        xml_str = export_ids_xml(reqs)
        root = ET.fromstring(xml_str)

        ns = {"ids": "http://standards.buildingsmart.org/IDS"}
        attrs = root.findall(".//ids:attribute", ns)
        assert len(attrs) == 1

    def test_export_with_min_max(self) -> None:
        """Numeric range constraints are exported as xs:restriction."""
        import xml.etree.ElementTree as ET

        reqs = [
            UniversalRequirement(
                element_filter={"ifc_class": "IFCWALL"},
                property_group="Pset_WallCommon",
                property_name="Width",
                constraint_def={"datatype": "IFCREAL", "cardinality": "required", "min": 0.1, "max": 0.5},
                context={"ifc_version": "IFC4", "use_case": "Test"},
            ),
        ]
        xml_str = export_ids_xml(reqs)

        xs_ns = "http://www.w3.org/2001/XMLSchema"
        root = ET.fromstring(xml_str)
        min_els = root.findall(f".//{{{xs_ns}}}minInclusive")
        max_els = root.findall(f".//{{{xs_ns}}}maxInclusive")
        assert len(min_els) == 1
        assert min_els[0].get("value") == "0.1"
        assert len(max_els) == 1
        assert max_els[0].get("value") == "0.5"
