"""End-to-end CSV / Excel formula-injection regression tests.

Reproduces the BUG-CSV-INJECTION attack: a user who can edit a BOQ position
(e.g. via the import-from-Excel endpoint, or a malicious tenant insider)
stores a description such as ``=cmd|'/c calc'!A0``. The string sits
harmlessly in PostgreSQL; the danger arrives when somebody else
*downloads the export* — Excel will interpret the leading ``=`` as a
formula and execute the payload.

The fix is **output-side neutralisation**: every string cell that contains
user data is run through :func:`app.core.csv_safety.neutralise_formula`,
which prepends a single apostrophe to anything starting with one of
``= + - @ \\t \\r``.

These tests verify that:
    1. The CSV bytes contain ``'=cmd|...`` rather than ``=cmd|...``
    2. The XLSX cell value, read back via ``openpyxl``, also contains the
       prepended apostrophe.
    3. Vanilla descriptions (``Concrete C30/37``) are byte-for-byte
       unchanged — the defence must not corrupt legitimate content.

Run::

    cd backend
    python -m pytest tests/integration/test_export_formula_safety.py -v --tb=short
"""

from __future__ import annotations

import asyncio
import io
import uuid

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient

from app.main import create_app

# Canonical CSV-injection payload from the OWASP cheat sheet. Excel parses
# the leading ``=`` as a formula and the ``cmd|...`` invokes DDE / RCE.
ATTACK_PAYLOAD = "=cmd|'/c calc'!A0"
PLUS_PAYLOAD = "+1+1+cmd|' /c calc'!A0"
MINUS_PAYLOAD = "-2+3+cmd|' /c calc'!A0"
AT_PAYLOAD = "@SUM(1+1)*cmd|' /c calc'!A0"
SAFE_DESCRIPTION = "Concrete C30/37 — foundation"


# ── Module-scoped fixtures (mirrors test_boq_regression.py to share style) ──


@pytest_asyncio.fixture(scope="module")
async def shared_client():
    app = create_app()

    from contextlib import asynccontextmanager

    @asynccontextmanager
    async def lifespan_ctx():
        async with app.router.lifespan_context(app):
            yield

    async with lifespan_ctx():
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as ac:
            yield ac


@pytest_asyncio.fixture(scope="module")
async def shared_auth(shared_client: AsyncClient) -> dict[str, str]:
    unique = uuid.uuid4().hex[:8]
    email = f"csvinj-{unique}@test.io"
    password = f"CsvInjection{unique}9"

    reg = await shared_client.post(
        "/api/v1/users/auth/register",
        json={
            "email": email,
            "password": password,
            "full_name": "CSV Injection Tester",
            "role": "admin",
        },
    )
    assert reg.status_code == 201, f"Registration failed: {reg.text}"

    # The public /register endpoint demotes new accounts to ``viewer`` for
    # security (BUG-327/386); promote via direct DB write so the test user
    # has the ``boq.create`` permission needed below.
    from ._auth_helpers import promote_to_admin

    await promote_to_admin(email)

    token = ""
    for attempt in range(3):
        resp = await shared_client.post(
            "/api/v1/users/auth/login",
            json={"email": email, "password": password},
        )
        data = resp.json()
        token = data.get("access_token", "")
        if token:
            break
        if "Too many login attempts" in data.get("detail", ""):
            await asyncio.sleep(5 * (attempt + 1))
            continue
        break
    assert token, f"Login failed: {data}"
    return {"Authorization": f"Bearer {token}"}


# ── BOQ scaffolding helper ─────────────────────────────────────────────────


async def _create_boq_with_positions(
    client: AsyncClient,
    auth: dict[str, str],
    descriptions: list[str],
) -> str:
    """Create a project + BOQ + positions and return the BOQ id.

    Each entry in *descriptions* becomes the description of one position
    so we can craft one row per attack vector we want to verify.
    """
    resp = await client.post(
        "/api/v1/projects/",
        json={
            "name": f"CSV Injection Project {uuid.uuid4().hex[:6]}",
            "description": "Project to verify CSV/XLSX formula neutralisation",
            "region": "DACH",
            "classification_standard": "din276",
            "currency": "EUR",
        },
        headers=auth,
    )
    assert resp.status_code == 201, resp.text
    project_id = resp.json()["id"]

    resp = await client.post(
        "/api/v1/boq/boqs/",
        json={
            "project_id": project_id,
            "name": "CSV Injection BOQ",
            "description": "BOQ used to assert formula neutralisation on export",
        },
        headers=auth,
    )
    assert resp.status_code == 201, resp.text
    boq_id = resp.json()["id"]

    for idx, desc in enumerate(descriptions, start=1):
        # Trailing slash matters: the endpoint is registered at
        # ``/positions/`` and FastAPI's redirect-on-mismatch is disabled
        # in this app, so a bare ``/positions`` returns 404.
        resp = await client.post(
            f"/api/v1/boq/boqs/{boq_id}/positions/",
            json={
                "boq_id": boq_id,
                "ordinal": f"01.{idx:03d}",
                "description": desc,
                "unit": "m3",
                "quantity": 1.0,
                "unit_rate": 100.0,
            },
            headers=auth,
        )
        assert resp.status_code == 201, f"Position create failed: {resp.text}"

    return boq_id


# ── CSV export ─────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_csv_export_neutralises_formula_payload(
    shared_client: AsyncClient,
    shared_auth: dict[str, str],
) -> None:
    """The raw CSV bytes must contain the apostrophe-prefixed payload, never
    the bare formula. Anything else lets Excel execute attacker code on the
    user's workstation when they open the downloaded file."""
    boq_id = await _create_boq_with_positions(
        shared_client,
        shared_auth,
        [ATTACK_PAYLOAD, PLUS_PAYLOAD, MINUS_PAYLOAD, AT_PAYLOAD, SAFE_DESCRIPTION],
    )

    resp = await shared_client.get(
        f"/api/v1/boq/boqs/{boq_id}/export/csv/", headers=shared_auth
    )
    assert resp.status_code == 200, resp.text
    body = resp.text

    # Each dangerous payload appears, but only with the leading apostrophe.
    # ``"'=cmd|..."`` is what we want; the unescaped ``"=cmd|..."`` is the
    # vulnerability. csv.writer wraps strings starting with ``'`` in quotes,
    # so we look for the quoted form.
    for payload in (ATTACK_PAYLOAD, PLUS_PAYLOAD, MINUS_PAYLOAD, AT_PAYLOAD):
        neutralised = "'" + payload
        assert neutralised in body, (
            f"Expected neutralised payload {neutralised!r} in CSV body"
        )

    # Sanity: legitimate text is not mangled — no spurious apostrophe.
    assert SAFE_DESCRIPTION in body
    assert ("'" + SAFE_DESCRIPTION) not in body


# ── Excel export ───────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_xlsx_export_neutralises_formula_payload(
    shared_client: AsyncClient,
    shared_auth: dict[str, str],
) -> None:
    """openpyxl reads the cell back as plain text; the apostrophe is part of
    the *stored* string value, which is exactly how Excel signals "this is
    text, not a formula" when the file is opened."""
    from openpyxl import load_workbook

    boq_id = await _create_boq_with_positions(
        shared_client,
        shared_auth,
        [ATTACK_PAYLOAD, PLUS_PAYLOAD, MINUS_PAYLOAD, AT_PAYLOAD, SAFE_DESCRIPTION],
    )

    resp = await shared_client.get(
        f"/api/v1/boq/boqs/{boq_id}/export/excel/", headers=shared_auth
    )
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers.get("content-type", "")

    wb = load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb.active

    # Collect every string value from the sheet — descriptions live in column
    # 2 (Description) but we don't pin the row index because subtotal rows
    # may shift it. Iterating the whole sheet is safer and still cheap.
    cells: list[str] = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str):
                cells.append(v)

    for payload in (ATTACK_PAYLOAD, PLUS_PAYLOAD, MINUS_PAYLOAD, AT_PAYLOAD):
        neutralised = "'" + payload
        assert neutralised in cells, (
            f"Expected neutralised payload {neutralised!r} in XLSX cells; "
            f"got {cells!r}"
        )
        # The bare payload must NEVER appear — that's the vulnerability.
        assert payload not in cells, (
            f"Bare payload {payload!r} leaked into XLSX cells: {cells!r}"
        )

    # Sanity: legitimate text round-trips unchanged.
    assert SAFE_DESCRIPTION in cells


@pytest.mark.asyncio
async def test_csv_export_does_not_modify_safe_descriptions(
    shared_client: AsyncClient,
    shared_auth: dict[str, str],
) -> None:
    """Belt-and-braces regression: if neutralisation ever started padding
    benign content, BOQ round-trips would silently corrupt every legitimate
    description with a leading apostrophe."""
    boq_id = await _create_boq_with_positions(
        shared_client,
        shared_auth,
        ["Concrete C30/37", "Formwork for foundations", "Reinforcing steel BSt 500 S"],
    )

    resp = await shared_client.get(
        f"/api/v1/boq/boqs/{boq_id}/export/csv/", headers=shared_auth
    )
    assert resp.status_code == 200, resp.text
    body = resp.text

    for desc in ("Concrete C30/37", "Formwork for foundations", "Reinforcing steel BSt 500 S"):
        assert desc in body
        assert ("'" + desc) not in body
