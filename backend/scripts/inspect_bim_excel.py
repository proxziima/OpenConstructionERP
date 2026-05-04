"""‌⁠‍Inspect a DDC cad2data / RvtExporter Excel (or Parquet) output.

Run this against the ``original.xlsx`` that DDC produces next to a saved
``geometry.dae`` to diagnose why BIM elements come back without ``mesh_ref``,
``storey``, or ``bounding_box``.

What it does
------------
1. Loads every header cell from the first sheet.
2. Normalises each header the same way the BIM uploader does
   (:func:`app.modules.bim_hub.router._match_bim_column`) and reports which
   canonical field each header maps to.
3. Prints the first three data rows so you can eyeball the content.
4. Tells you, for the current ``_BIM_COLUMN_ALIASES`` dict, which columns
   would populate ``mesh_ref`` / ``storey`` / ``bounding_box`` and whether
   any of them are missing.

Usage::

    cd backend
    python -m scripts.inspect_bim_excel path/to/original.xlsx
    # or
    python scripts/inspect_bim_excel.py path/to/elements.parquet
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any


def _load_excel(path: Path) -> tuple[list[str], list[list[Any]]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise SystemExit(f"{path} has no active worksheet")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None) or ()
    headers = [str(h or "").strip() for h in raw_headers]

    data_rows: list[list[Any]] = []
    for row in rows_iter:
        data_rows.append(list(row))
        if len(data_rows) >= 20:
            break

    wb.close()
    return headers, data_rows


def _load_parquet(path: Path) -> tuple[list[str], list[list[Any]]]:
    try:
        import pyarrow.parquet as pq  # type: ignore[import-not-found]
    except ImportError as exc:
        raise SystemExit("pyarrow is required to inspect parquet files") from exc

    table = pq.read_table(path)
    headers = list(table.column_names)
    rows = table.slice(0, 20).to_pylist()
    data_rows = [[r.get(h) for h in headers] for r in rows]
    return headers, data_rows


def main() -> int:
    if len(sys.argv) != 2:
        print(__doc__)
        return 2

    path = Path(sys.argv[1]).expanduser().resolve()
    if not path.is_file():
        print(f"ERROR: file not found: {path}")
        return 1

    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        headers, data_rows = _load_excel(path)
    elif suffix == ".parquet":
        headers, data_rows = _load_parquet(path)
    else:
        print(f"ERROR: unsupported file type: {suffix}")
        return 1

    print(f"FILE: {path}")
    print(f"COLUMNS: {len(headers)}")
    print()

    # Use the live alias dict from the uploader so this script catches drift.
    try:
        from app.modules.bim_hub.router import _BIM_COLUMN_ALIASES, _match_bim_column
    except ImportError:
        # Fallback: add backend/ to sys.path and retry.
        backend_dir = Path(__file__).resolve().parent.parent
        sys.path.insert(0, str(backend_dir))
        from app.modules.bim_hub.router import _BIM_COLUMN_ALIASES, _match_bim_column

    # Strip DDC type suffixes so headers like "Level : String" match aliases.
    def _clean(h: str) -> str:
        return h.split(" : ")[0].strip() if " : " in h else h

    print("HEADER -> CANONICAL")
    print("-" * 60)
    canonical_hits: dict[str, list[str]] = {}
    for h in headers:
        if not h:
            continue
        canonical = _match_bim_column(_clean(h))
        canonical_hits.setdefault(canonical or "", []).append(h)
        if canonical and canonical in _BIM_COLUMN_ALIASES:
            print(f"  {h!r:45s} -> {canonical}")

    print()
    print("CANONICAL COVERAGE (fields we care about)")
    print("-" * 60)
    fields_of_interest = (
        ("element_id", "stable_id / mesh_ref source"),
        ("mesh_ref", "explicit DAE node reference"),
        ("storey", "building level / floor"),
        ("bounding_box", "axis-aligned bounding box (pre-built JSON)"),
        ("bbox_min_x", "bbox min X (one of six numeric bbox columns)"),
        ("bbox_max_x", "bbox max X"),
    )
    for canonical, desc in fields_of_interest:
        matches = canonical_hits.get(canonical, [])
        mark = "OK   " if matches else "MISS "
        print(f"  {mark}{canonical:15s} {desc}")
        for m in matches:
            print(f"         found in: {m!r}")

    print()
    print("FIRST 3 DATA ROWS (non-empty cells only)")
    print("-" * 60)
    for i, row in enumerate(data_rows[:3]):
        print(f"-- row {i + 1} --")
        for h, v in zip(headers, row, strict=False):
            if v is None or v == "":
                continue
            text = str(v)
            if len(text) > 80:
                text = text[:77] + "..."
            print(f"  {h}: {text}")

    # Final verdict.
    missing = [c for c, _ in fields_of_interest if not canonical_hits.get(c)]
    print()
    if missing:
        print("VERDICT: MISSING columns for:", ", ".join(missing))
        print("  -> the uploader will emit NULL for these fields unless they")
        print("     are derivable from another column (e.g. mesh_ref from ID).")
    else:
        print("VERDICT: all tracked fields have at least one mapped column.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
