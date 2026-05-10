# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""Excel BoQ ingest for the match-elements module.

Parses a multi-language BoQ spreadsheet into the row-dict shape that
:class:`BoqAdapter` already understands. Mirrors the column-alias logic
in ``boq/router.py`` but extends the alias map across the global locale
spread (CJK / Slavic / Romance / Turkic / Arabic) because /match-elements
accepts arbitrary regional BoQ files — a German GAEB export, a Russian
смета, a Turkish keşif, all in a single workspace.

Public surface:

* :func:`parse_boq_xlsx` — bytes → list[dict] ready for SessionCreate.

The module imports :mod:`openpyxl` lazily so ``backend/app/__init__``
stays import-safe when the optional Excel toolchain is missing on a
slim deploy. The dependency is listed under the platform's standard
extras and is already a transitive dep of the BOQ module.
"""

from __future__ import annotations

import io
import logging
from typing import Any

logger = logging.getLogger(__name__)


# Multi-language column aliases. Compared lowercased + whitespace-stripped
# against the first row of the worksheet. Order is irrelevant — first
# alias-list-hit wins. Extending: add a header spelling to the relevant
# canonical list; never remove existing entries (callers may have shipped
# spreadsheets relying on them).
_ALIASES: dict[str, list[str]] = {
    "description": [
        # English
        "description", "desc", "text", "item", "item description",
        "work description", "scope", "scope of work",
        # German
        "beschreibung", "bezeichnung", "leistung", "leistungstext",
        "kurztext", "langtext",
        # Russian / Ukrainian / Belarusian
        "описание", "наименование", "наименование работ",
        "работы", "элемент", "состав работ", "опис",
        # Romance
        "descripción", "descripcion", "descrição", "descricao",
        "descrizione", "désignation", "designation", "obra",
        # Turkic / Slavic
        "açıklama", "iş tanımı", "opis", "opis prac",
        # CJK
        "説明", "工事内容", "描述", "项目描述", "설명", "내역",
        # Arabic
        "الوصف", "البند",
    ],
    "qty": [
        "qty", "quantity", "amount", "qty.", "quantity (qty)",
        # German / Dutch
        "menge", "anzahl", "hoeveelheid",
        # Russian
        "количество", "кол-во", "кол.",
        # Romance
        "cantidad", "quantidade", "quantité", "quantita", "quantità",
        # Turkic / Slavic
        "miktar", "ilość",
        # CJK
        "数量", "수량",
        # Arabic
        "الكمية",
    ],
    "unit": [
        "unit", "uom", "unit of measure", "unit_of_measure",
        # German
        "einheit", "me", "mengeneinheit",
        # Russian
        "ед", "ед.изм", "ед.изм.", "единица", "ед изм",
        # Romance
        "unidad", "unidade", "unité", "unita", "unità",
        # Turkic / Slavic
        "birim", "jednostka",
        # CJK
        "単位", "单位", "단위",
        # Arabic
        "الوحدة",
    ],
    "code": [
        "code", "rate code", "rate_code",
        # German
        "kennzeichen", "pos.kennz", "ord.-nr", "art.-nr",
        # Russian
        "код", "артикул", "шифр",
        # Romance
        "código", "codigo", "codice",
        # CJK / Arabic / Turkic
        "コード", "代码", "코드", "kod", "الكود",
    ],
    "category": [
        "category", "section", "trade",
        # German
        "kategorie", "abteilung", "abschnitt", "gewerk",
        # Russian
        "категория", "раздел", "вид",
        # Romance
        "categoría", "categoria", "sección", "secção", "categoría",
        "section", "categoria", "rubrique",
        # CJK
        "カテゴリー", "类别", "분류",
    ],
    "source_lang": [
        "source_lang", "lang", "language", "language code",
        "язык",
    ],
}


def _match_column(header: Any) -> str | None:
    """Return the canonical column name for a header cell, or ``None``.

    Lookup is whitespace-insensitive, lowercased, and tolerant of header
    cells that arrive as a non-string (openpyxl returns ``None`` for
    blank cells, ``int``/``datetime`` for the rare misformatted header).
    """
    if header is None:
        return None
    raw = str(header).strip().lower()
    if not raw:
        return None
    for canonical, aliases in _ALIASES.items():
        if raw == canonical or raw in aliases:
            return canonical
    return None


def _to_float_qty(value: Any) -> float | None:
    """Coerce a qty cell to float, tolerating ``"1.234,56"`` Euro form.

    Returns ``None`` for blanks, NaN, and non-numeric strings — the
    caller drops those rows so downstream matchers don't divide by zero.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        f = float(value)
        return f if f == f else None
    s = str(value).strip()
    if not s:
        return None
    if "," in s and "." in s:
        last_c = s.rfind(",")
        last_d = s.rfind(".")
        if last_c > last_d:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_boq_xlsx(content: bytes) -> list[dict[str, Any]]:
    """Parse an xlsx upload into a list of BoQ row dicts.

    Reads the first (active) worksheet, treats the first row as headers,
    and extracts cells matching the alias map. Returns rows ready to
    feed into :class:`SessionCreate.boq_rows` and from there to
    :class:`BoqAdapter`.

    Rules:

    * Rows without a ``description`` cell are skipped (the matcher needs
      at least the freeform text — without it neither dense, sparse,
      nor exact-code paths can land on a CWICR rate).
    * ``qty`` is parsed numerically; non-numeric values are dropped from
      the row (the BoqAdapter then defaults to ``count=1.0``).
    * Other recognised columns pass through as strings (trimmed).
    * Unknown columns are ignored — they don't pollute the dict so
      downstream group-by chips stay clean.

    Args:
        content: The xlsx file bytes as uploaded.

    Returns:
        ``[{description, qty?, unit?, code?, category?, source_lang?}, ...]``

    Raises:
        ValueError: When the workbook can't be opened, has no
            worksheets, or has no header row.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(
            f"Could not open the uploaded file as an xlsx workbook: {exc}"
        ) from exc

    try:
        ws = wb.active
        if ws is None:
            raise ValueError("Workbook has no active worksheet.")

        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            raise ValueError("Workbook is empty (no header row).")

        column_map: dict[int, str] = {}
        for idx, hdr in enumerate(headers):
            canon = _match_column(hdr)
            if canon:
                column_map[idx] = canon

        if "description" not in column_map.values():
            raise ValueError(
                "No 'Description' column detected. Add a header named "
                "'Description' (or its localised equivalent — 'Beschreibung', "
                "'Описание', 'Descripción', '描述', etc.)."
            )

        out: list[dict[str, Any]] = []
        for row in rows_iter:
            entry: dict[str, Any] = {}
            for idx, val in enumerate(row):
                key = column_map.get(idx)
                if key is None or val is None:
                    continue
                if key == "qty":
                    f = _to_float_qty(val)
                    if f is not None:
                        entry["qty"] = f
                else:
                    entry[key] = (
                        str(val).strip() if isinstance(val, str) else val
                    )
            description = entry.get("description")
            if isinstance(description, str) and description.strip():
                out.append(entry)
        return out
    finally:
        wb.close()


__all__ = ["parse_boq_xlsx"]
