"""Excel (.xlsx) and CSV BOQ importer.

Generic spreadsheet ingester with three classification heuristics on top
of the column-alias mapper:

* **NRM** — UK New Rules of Measurement. Detects element codes like
  ``2.6.1`` and section headers like ``Element 2 — Substructure``.
* **MasterFormat** — US CSI MasterFormat. Detects 6-digit codes like
  ``03 30 00`` and division headers like
  ``Division 03 — Cast-in-Place Concrete``.
* **Generic** — anything else goes into ``classification["code"]``.

Epic I3 (refactor) keeps the parser pure (no DB I/O, no FastAPI types);
all the per-row error reporting, dry-run handling and inline validation
the route used to do inline now live in the dispatcher route. Epics
I9 / I10 wire in the NRM and MasterFormat division detectors as
:func:`_infer_classification`.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from typing import Any, ClassVar, Literal

from app.core.file_signature import detect as detect_signature
from app.modules.boq.importers._base import (
    BOQImporter,
    ImportedBOQ,
    ImportedPosition,
    ImporterParseError,
)
from app.modules.boq.importers._encoding import (
    decode_text_bytes,
    parse_numeric_cell,
    safe_float,
)

logger = logging.getLogger(__name__)


# ── Column alias map (mirrors the historic ``_COLUMN_ALIASES``) ────────────
#
# Canonical column → set of accepted header strings (lowercased). Editing
# this map is the supported extension point for new locale variants
# (Polish ``Ilosc``, Italian ``Quantità`` etc. land here).
_COLUMN_ALIASES: dict[str, frozenset[str]] = {
    "ordinal": frozenset(
        {
            "pos",
            "pos.",
            "position",
            "ordinal",
            "nr",
            "nr.",
            "no",
            "no.",
            "ord",
            "item",
            "item no",
            "ref",
            "code",
        }
    ),
    "description": frozenset(
        {
            "description",
            "desc",
            "text",
            "beschreibung",
            "leistung",
            "designación",
            "designacion",
            "designation",
            "désignation",
            "descripción",
            "descripcion",
            "descrizione",
            "opis",
            "наименование",
        }
    ),
    "unit": frozenset(
        {
            "unit",
            "einheit",
            "me",
            "uds",
            "ud",
            "u",
            "unité",
            "unidad",
            "unità",
            "jed",
            "ед",
            "ед.",
        }
    ),
    "quantity": frozenset(
        {
            "quantity",
            "qty",
            "menge",
            "cantidad",
            "cant",
            "cant.",
            "quantité",
            "quantita",
            "quantità",
            "ilość",
            "ilosc",
            "количество",
            "кол-во",
        }
    ),
    "unit_rate": frozenset(
        {
            "unit rate",
            "rate",
            "unitrate",
            "ep",
            "einheitspreis",
            "preis",
            "precio",
            "prezzo",
            "prix",
            "цена",
        }
    ),
    "total": frozenset(
        {"total", "amount", "gesamt", "gesamtpreis", "importe", "subtotal", "стоимость"}
    ),
    "classification": frozenset(
        {
            "classification",
            "din 276",
            "din276",
            "kg",
            "nrm",
            "code",
            "csi",
            "masterformat",
            "element",
            "division",
            "category",
            "trade",
        }
    ),
}


def _match_column(header: str) -> str | None:
    """Match a header string to a canonical column name using the alias map."""
    normalised = header.strip().lower()
    for canonical, aliases in _COLUMN_ALIASES.items():
        if normalised in aliases:
            return canonical
    return None


def _detect_file_format(content_head: bytes) -> Literal["xlsx", "csv", "parquet", "unknown"]:
    """Identify an upload by its magic bytes (BUG-UPLOAD01 from the legacy code).

    A ``.exe`` renamed to ``.xlsx`` would otherwise be handed to
    ``openpyxl`` — best case a parse exception, worst case the bytes
    land in our buffers + logs before we error.
    """
    if not content_head:
        return "unknown"
    sig = detect_signature(content_head)
    if sig == "zip":  # XLSX = OOXML zip
        return "xlsx"
    if content_head[:4] == b"PAR1":
        return "parquet"
    if b"\x00" in content_head:
        return "unknown"
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            decoded = content_head.decode(encoding)
        except UnicodeDecodeError:
            continue
        if any(sep in decoded for sep in (",", ";", "\t", "|", "\n")):
            return "csv"
    return "unknown"


# ── Classification heuristics (Epics I9 + I10) ─────────────────────────────


# NRM codes: ``N.N.N`` or ``N.N`` (e.g. ``2.6.1``, ``2.6``). NRM 1 / NRM 2
# tops out at four levels but two/three are the common case in tender
# documents.
_NRM_CODE_RE = re.compile(r"^(\d{1,2}\.){1,3}\d{1,2}$")

# NRM element header text e.g. ``"Element 2 — Substructure"``,
# ``"Group element 2.6 — External walls"``.
_NRM_HEADER_RE = re.compile(
    r"^(group\s+)?element\s+(\d{1,2}(?:\.\d{1,2})*)\b", re.IGNORECASE
)

# MasterFormat: ``XX XX XX`` or ``XX.XX.XX`` or ``XX-XX-XX`` (2-2-2 digits).
# Sub-codes ``XX XX XX.XX`` are allowed.
_MASTERFORMAT_CODE_RE = re.compile(
    r"^(\d{2})[\s.\-](\d{2})[\s.\-](\d{2})(?:\.(\d{2}))?$"
)

# MasterFormat division header text e.g. ``"Division 03 — Concrete"``,
# ``"03 30 00 Cast-in-Place Concrete"``.
_MASTERFORMAT_HEADER_RE = re.compile(
    r"^division\s+(\d{2})\b", re.IGNORECASE
)


def _infer_classification(
    code_text: str,
    description: str,
) -> dict[str, Any]:
    """Heuristic classification from a raw code cell + description.

    Tries NRM and MasterFormat patterns; anything else falls through to
    ``{"code": code_text}`` (the historic generic behaviour).
    """
    code = code_text.strip()
    desc = (description or "").strip()
    classification: dict[str, Any] = {}

    # NRM element header in the description ("Element 2 — Substructure").
    m = _NRM_HEADER_RE.match(desc)
    if m:
        classification["nrm"] = m.group(2)
    # NRM code pattern in the code cell ("2.6.1").
    if code and _NRM_CODE_RE.match(code):
        classification["nrm"] = code

    # MasterFormat 6-digit code in the code cell ("03 30 00").
    m = _MASTERFORMAT_CODE_RE.match(code) if code else None
    if m:
        # Normalise to spaced form "XX XX XX[.XX]".
        parts = [m.group(1), m.group(2), m.group(3)]
        mf = " ".join(parts)
        if m.group(4):
            mf = f"{mf}.{m.group(4)}"
        classification["masterformat"] = mf

    # MasterFormat division header in the description ("Division 03 —").
    m = _MASTERFORMAT_HEADER_RE.match(desc)
    if m:
        # Pad to canonical 6-digit form for downstream rules.
        div = m.group(1)
        # If the description contains a fuller code further along, keep it,
        # else stub the level-2 + level-3 to ``00``.
        if "masterformat" not in classification:
            classification["masterformat"] = f"{div} 00 00"

    # Fallback: stash the raw code so the editor can show it. Skip if we
    # already mapped it to a structured field above.
    if code and "nrm" not in classification and "masterformat" not in classification:
        classification["code"] = code

    return classification


# ── Row parsing helpers ─────────────────────────────────────────────────────


def _parse_rows_from_csv(content_bytes: bytes) -> list[dict[str, Any]]:
    """Decode + parse a CSV into a list of canonical-key dicts."""
    text, _ = decode_text_bytes(content_bytes)
    # Detect delimiter from the first 4 KB.
    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ImporterParseError("CSV file is empty or has no header row")

    column_map: dict[int, str] = {}
    raw_header_strings: list[str] = []
    for idx, hdr in enumerate(raw_headers):
        raw_header_strings.append(str(hdr or ""))
        canonical = _match_column(str(hdr or ""))
        if canonical:
            column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical:
                row[canonical] = val.strip() if isinstance(val, str) else val
        if row:
            rows.append(row)
    return rows


def _parse_rows_from_excel(
    content_bytes: bytes,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Read an .xlsx file's first worksheet into canonical-key dicts.

    Returns ``(rows, import_metadata)``; metadata preserves the raw
    column ordering so a later export can round-trip back to the
    user's original spreadsheet layout.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ImporterParseError("Excel file has no worksheets")

    sheet_names = wb.sheetnames

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        raise ImporterParseError("Excel file is empty or has no header row")

    original_columns = [str(h) if h is not None else "" for h in raw_headers]
    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        if hdr is not None:
            canonical = _match_column(str(hdr))
            if canonical:
                column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None:
                row[canonical] = val
        if row:
            rows.append(row)
    wb.close()

    import_metadata = {
        "original_columns": original_columns,
        "column_mapping": {str(k): v for k, v in column_map.items()},
        "sheet_names": sheet_names,
        "total_rows": len(rows),
    }
    return rows, import_metadata


_TOTAL_ROW_DESCRIPTIONS = {
    "grand total",
    "total",
    "summe",
    "gesamt",
    "gesamtsumme",
    "subtotal",
    "zwischensumme",
}


_IMPORT_MAX_QUANTITY = 1e9
_IMPORT_MAX_UNIT_RATE = 1e8


def _rows_to_positions(
    rows: list[dict[str, Any]],
    *,
    source: str = "excel_import",
) -> ImportedBOQ:
    """Convert canonical rows into :class:`ImportedPosition` objects.

    Carries the sanity bounds + section-row + summary-row detection that
    the legacy inline parser used. Per-row errors are collected on the
    returned :class:`ImportedBOQ` rather than raised so the dispatcher
    can return them as a structured list.
    """
    result = ImportedBOQ(source_format="csv-or-xlsx")
    auto_ordinal = 1

    # Pre-compute a median unit rate across the file so we can warn on
    # any single position that's >10× above (likely a tampered export).
    rate_samples = sorted(
        v
        for v in (safe_float(r.get("unit_rate"), default=0.0) for r in rows)
        if v > 0
    )
    median_rate = rate_samples[len(rate_samples) // 2] if rate_samples else 0.0

    for row_idx, row in enumerate(rows, start=2):
        try:
            description = str(row.get("description", "")).strip()
            if not description:
                result.skipped += 1
                continue

            desc_lower = description.lower()
            if desc_lower in _TOTAL_ROW_DESCRIPTIONS:
                result.skipped += 1
                continue
            if desc_lower.startswith("subtotal:") or desc_lower.startswith("zwischensumme:"):
                result.skipped += 1
                continue

            ordinal = str(row.get("ordinal", "")).strip()
            if not ordinal:
                ordinal = str(auto_ordinal)
            auto_ordinal += 1

            unit_raw = str(row.get("unit", "")).strip()
            quantity_raw = row.get("quantity")
            unit_rate_raw = row.get("unit_rate")
            quantity, q_err = parse_numeric_cell(quantity_raw)
            unit_rate, r_err = parse_numeric_cell(unit_rate_raw)
            if q_err is not None:
                result.errors.append(
                    {
                        "row": row_idx,
                        "ordinal": ordinal,
                        "error": f"Invalid quantity at row {row_idx}: {q_err}",
                    }
                )
                continue
            if r_err is not None:
                result.errors.append(
                    {
                        "row": row_idx,
                        "ordinal": ordinal,
                        "error": f"Invalid unit_rate at row {row_idx}: {r_err}",
                    }
                )
                continue
            assert quantity is not None
            assert unit_rate is not None

            # Section detection: a row with a description but no unit /
            # quantity / rate is a section header from our own exporter.
            is_section_row = (
                not unit_raw
                and (quantity_raw in (None, "", 0, 0.0))
                and (unit_rate_raw in (None, "", 0, 0.0))
            )
            if is_section_row:
                result.positions.append(
                    ImportedPosition(
                        description=description,
                        ordinal=ordinal,
                        unit="section",
                        quantity=0.0,
                        unit_rate=0.0,
                        classification={},
                        source=source,
                        metadata={
                            "import_row_index": row_idx,
                            "section_header": True,
                        },
                        is_section=True,
                    )
                )
                continue

            unit = unit_raw or "pcs"

            # Range guards — reject obvious tamper / typo errors.
            if not (0 <= quantity <= _IMPORT_MAX_QUANTITY):
                result.errors.append(
                    {
                        "row": row_idx,
                        "ordinal": ordinal,
                        "error": f"Quantity out of range: {quantity}",
                    }
                )
                continue
            if not (0 <= unit_rate <= _IMPORT_MAX_UNIT_RATE):
                result.errors.append(
                    {
                        "row": row_idx,
                        "ordinal": ordinal,
                        "error": f"Unit rate out of range: {unit_rate}",
                    }
                )
                continue

            # Soft warnings.
            if median_rate > 0 and unit_rate > median_rate * 10:
                result.warnings.append(
                    {
                        "row": row_idx,
                        "ordinal": ordinal,
                        "severity": "warning",
                        "message": (
                            f"Unit rate {unit_rate:.2f} is >10× the file median "
                            f"({median_rate:.2f}) — possible typo or tampered export."
                        ),
                    }
                )
            if quantity == 0:
                result.warnings.append(
                    {
                        "row": row_idx,
                        "ordinal": ordinal,
                        "severity": "info",
                        "message": "Quantity is zero — position imported but contributes no cost.",
                    }
                )
            if unit_rate == 0:
                result.warnings.append(
                    {
                        "row": row_idx,
                        "ordinal": ordinal,
                        "severity": "info",
                        "message": "Unit rate is zero — position imported without a rate.",
                    }
                )

            # Heuristic classification (Epics I9 + I10).
            class_value = str(row.get("classification", "")).strip()
            classification = _infer_classification(class_value, description)

            result.positions.append(
                ImportedPosition(
                    description=description,
                    ordinal=ordinal,
                    unit=unit,
                    quantity=quantity,
                    unit_rate=unit_rate,
                    classification=classification,
                    source=source,
                    metadata={"import_row_index": row_idx},
                )
            )

        except Exception as exc:  # noqa: BLE001 — caller surfaces row #
            result.errors.append({"row": row_idx, "ordinal": "", "error": str(exc)})
            logger.warning("Excel/CSV row %d error: %s", row_idx, exc)

    return result


class ExcelImporter:
    """Generic Excel (.xlsx) / CSV importer with NRM + MasterFormat heuristics."""

    format_id: ClassVar[str] = "excel"
    extensions: ClassVar[tuple[str, ...]] = (".xlsx", ".csv")
    display_name: ClassVar[str] = "Excel / CSV BOQ"
    rule_packs: ClassVar[tuple[str, ...]] = ("boq_quality",)

    @classmethod
    def detect(cls, head_bytes: bytes, filename: str) -> bool:
        """Detect by magic bytes (xlsx zip header / CSV text) + extension."""
        if not head_bytes:
            return False
        name = filename.lower()
        if not any(name.endswith(ext) for ext in cls.extensions):
            return False
        fmt = _detect_file_format(head_bytes[:4096])
        if name.endswith(".xlsx"):
            return fmt == "xlsx"
        if name.endswith(".csv"):
            return fmt == "csv"
        return False

    @classmethod
    async def parse(cls, content: bytes, *, locale: str = "en") -> ImportedBOQ:
        """Parse an .xlsx or .csv BOQ into :class:`ImportedBOQ`."""
        if not content:
            raise ImporterParseError("Spreadsheet upload is empty")

        fmt = _detect_file_format(content[:4096])
        import_meta: dict[str, Any] = {}
        try:
            if fmt == "xlsx":
                rows, import_meta = _parse_rows_from_excel(content)
                source_format = "xlsx"
            elif fmt == "csv":
                rows = _parse_rows_from_csv(content)
                source_format = "csv"
            else:
                raise ImporterParseError(
                    f"Unsupported spreadsheet format: detected {fmt!r}"
                )
        except ImporterParseError:
            raise
        except Exception as exc:  # noqa: BLE001
            raise ImporterParseError(f"Could not parse spreadsheet: {exc}") from exc

        if not rows:
            raise ImporterParseError(
                "No data rows found. Check that the first row contains column headers."
            )

        result = _rows_to_positions(rows)
        result.source_format = source_format
        result.metadata = {
            "original_columns": import_meta.get("original_columns", []),
            "column_mapping": import_meta.get("column_mapping", {}),
            "sheet_names": import_meta.get("sheet_names", []),
            "total_rows_seen": len(rows),
        }
        return result
