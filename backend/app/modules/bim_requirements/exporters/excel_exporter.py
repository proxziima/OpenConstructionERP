"""‚Äå‚ÅÝ‚ÄçExcel exporter -- produces a formatted Excel workbook.

Generates a professional Excel file with:
- Sheet 1: Requirements table with formatted headers
- Sheet 2: Summary by element
- Sheet 3: Legend explaining columns and datatypes
"""

import io
import logging
from typing import Any

from app.modules.bim_requirements.parsers.base import UniversalRequirement

logger = logging.getLogger(__name__)

# Header definitions: (column letter, EN header, DE header, width)
_COLUMNS = [
    ("A", "Element / IFC Class", "Bauteil / IFC-Klasse", 25),
    ("B", "Classification", "Klassifikation", 20),
    ("C", "Property Group", "Merkmalsgruppe", 22),
    ("D", "Property Name", "Merkmal", 25),
    ("E", "Data Type", "Datentyp", 15),
    ("F", "Required", "Erforderlich", 13),
    ("G", "Allowed Values", "Zulaessige Werte", 30),
    ("H", "Min", "Min-Wert", 10),
    ("I", "Max", "Max-Wert", 10),
    ("J", "Unit", "Einheit", 10),
    ("K", "Pattern", "Muster (Regex)", 20),
    ("L", "Phase", "Leistungsphase", 15),
    ("M", "Actor", "Akteur", 18),
    ("N", "Use Case", "Anwendungsfall", 20),
    ("O", "Source", "Quelle", 18),
]

# Color constants (RGB hex without #)
_HEADER_BG = "1F3864"
_HEADER_FG = "FFFFFF"
_ALT_ROW_BG = "EBF5FB"
_REQUIRED_BG = "C6EFCE"
_OPTIONAL_BG = "FFEB9C"
_PROHIBITED_BG = "FFC7CE"


def export_excel(
    requirements: list[UniversalRequirement],
    title: str = "BIM Requirements",
    language: str = "en",
) -> bytes:
    """‚Äå‚ÅÝ‚ÄçExport requirements to a formatted Excel workbook.

    Args:
        requirements: List of universal requirements to export.
        title: Title shown on the Summary sheet.
        language: 'en' or 'de' for header language.

    Returns:
        Excel file content as bytes.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ‚îÄ‚îÄ Sheet 1: Requirements ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    ws = wb.active
    ws.title = "Requirements"  # type: ignore[union-attr]

    header_fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    header_font = Font(bold=True, color=_HEADER_FG, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill = PatternFill(start_color=_ALT_ROW_BG, end_color=_ALT_ROW_BG, fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    req_fill = PatternFill(start_color=_REQUIRED_BG, end_color=_REQUIRED_BG, fill_type="solid")
    opt_fill = PatternFill(start_color=_OPTIONAL_BG, end_color=_OPTIONAL_BG, fill_type="solid")
    prohib_fill = PatternFill(
        start_color=_PROHIBITED_BG, end_color=_PROHIBITED_BG, fill_type="solid"
    )

    # Write headers
    for col_idx, (_, en_name, de_name, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)  # type: ignore[union-attr]
        cell.value = en_name if language == "en" else de_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width  # type: ignore[union-attr]

    # Set row height for header
    ws.row_dimensions[1].height = 22  # type: ignore[union-attr]

    # Freeze top row and first column
    ws.freeze_panes = "B2"  # type: ignore[union-attr]

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}1"  # type: ignore[union-attr]

    # Write data rows
    for row_idx, req in enumerate(requirements, start=2):
        ef = req.element_filter or {}
        cd = req.constraint_def or {}
        ctx = req.context or {}

        row_data = [
            ef.get("ifc_class", ""),
            _format_classification(ef.get("classification")),
            req.property_group or "",
            req.property_name,
            cd.get("datatype", ""),
            cd.get("cardinality", ""),
            "; ".join(cd["enum"]) if "enum" in cd else "",
            cd.get("min", ""),
            cd.get("max", ""),
            cd.get("unit", ""),
            cd.get("pattern", ""),
            ctx.get("phase", ""),
            ctx.get("actor", ""),
            ctx.get("use_case", ""),
            ctx.get("source", ""),
        ]

        is_alt = row_idx % 2 == 0
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)  # type: ignore[union-attr]
            cell.value = str(value) if value != "" else ""
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill

        # Conditional formatting for cardinality column (F)
        card_cell = ws.cell(row=row_idx, column=6)  # type: ignore[union-attr]
        card_val = str(card_cell.value).lower() if card_cell.value else ""
        if card_val == "required":
            card_cell.fill = req_fill
        elif card_val == "optional":
            card_cell.fill = opt_fill
        elif card_val == "prohibited":
            card_cell.fill = prohib_fill

    # ‚îÄ‚îÄ Sheet 2: Summary ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    ws_summary = wb.create_sheet("Summary")
    ws_summary.cell(row=1, column=1).value = "Summary"
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_summary.cell(row=2, column=1).value = title

    # Element summary
    element_counts: dict[str, int] = {}
    for req in requirements:
        ifc_class = (req.element_filter or {}).get("ifc_class", "Unspecified")
        element_counts[ifc_class] = element_counts.get(ifc_class, 0) + 1

    ws_summary.cell(row=4, column=1).value = "Element"
    ws_summary.cell(row=4, column=1).font = Font(bold=True)
    ws_summary.cell(row=4, column=2).value = "Requirements Count"
    ws_summary.cell(row=4, column=2).font = Font(bold=True)

    for i, (element, count) in enumerate(
        sorted(element_counts.items()), start=5
    ):
        ws_summary.cell(row=i, column=1).value = element
        ws_summary.cell(row=i, column=2).value = count

    ws_summary.cell(row=len(element_counts) + 6, column=1).value = "Total"
    ws_summary.cell(row=len(element_counts) + 6, column=1).font = Font(bold=True)
    ws_summary.cell(row=len(element_counts) + 6, column=2).value = len(requirements)
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 20

    # ‚îÄ‚îÄ Sheet 3: Legend ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ
    ws_legend = wb.create_sheet("Legend")
    ws_legend.cell(row=1, column=1).value = "Legend"
    ws_legend.cell(row=1, column=1).font = Font(bold=True, size=14)

    legend_rows = [
        ("Column", "Description"),
        ("Element / IFC Class", "IFC entity class (e.g. IFCWALL, IFCBEAM)"),
        ("Classification", "Classification system and code (e.g. Uniclass 2015: EF_25_10)"),
        ("Property Group", "Property set name (e.g. Pset_WallCommon)"),
        ("Property Name", "Name of the required property or attribute"),
        ("Data Type", "IFC data type (IFCLABEL, IFCREAL, IFCBOOLEAN, etc.)"),
        ("Required", "Cardinality: required, optional, or prohibited"),
        ("Allowed Values", "Semicolon-separated list of permitted values"),
        ("Min / Max", "Numeric range boundaries"),
        ("Unit", "Unit of measurement (m, m2, kg, etc.)"),
        ("Pattern", "Regular expression pattern for value validation"),
        ("Phase", "Project phase or milestone (LP3, RIBA Stage 3, etc.)"),
        ("Actor", "Responsible party (Architect, Structural Engineer, etc.)"),
        ("Use Case", "Purpose or use case for the requirement"),
        ("Source", "Source standard or document"),
    ]

    for i, (col_name, desc) in enumerate(legend_rows, start=3):
        ws_legend.cell(row=i, column=1).value = col_name
        ws_legend.cell(row=i, column=2).value = desc
        if i == 3:
            ws_legend.cell(row=i, column=1).font = Font(bold=True)
            ws_legend.cell(row=i, column=2).font = Font(bold=True)

    ws_legend.column_dimensions["A"].width = 25
    ws_legend.column_dimensions["B"].width = 60

    # Stamp document properties so downloaded XLSX carries our authorship.
    try:
        wb.properties.creator = "OpenConstructionERP ¬∑ DataDrivenConstruction"
        wb.properties.lastModifiedBy = "OpenConstructionERP"
        wb.properties.description = (
            "Generated by OpenConstructionERP "
            "(https://openconstructionerp.com) ¬∑ DDC-CWICR-OE-2026"
        )
        wb.properties.keywords = "DDC-CWICR-OE,OpenConstructionERP,BIM Requirements"
    except Exception:  # noqa: BLE001 ‚Äî best-effort metadata stamp
        pass

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _format_classification(classif: Any) -> str:
    """‚Äå‚ÅÝ‚ÄçFormat a classification dict as a display string."""
    if not classif:
        return ""
    if isinstance(classif, dict):
        system = classif.get("system", "")
        value = classif.get("value", "")
        if system and value:
            return f"{system}: {value}"
        return value or system
    return str(classif)


def generate_template() -> bytes:
    """Generate a downloadable Excel template with headers and example rows.

    Returns:
        Excel file content as bytes.
    """
    example_reqs = [
        UniversalRequirement(
            element_filter={"ifc_class": "IFCWALL"},
            property_group="Pset_WallCommon",
            property_name="FireRating",
            constraint_def={
                "datatype": "IFCLABEL",
                "cardinality": "required",
                "enum": ["REI60", "REI90", "REI120"],
            },
            context={"phase": "LP3", "actor": "Architect", "use_case": "Fire Safety"},
        ),
        UniversalRequirement(
            element_filter={"ifc_class": "IFCWALL"},
            property_group="Pset_WallCommon",
            property_name="ThermalTransmittance",
            constraint_def={
                "datatype": "IFCREAL",
                "cardinality": "required",
                "max": 0.28,
                "unit": "W/(m2K)",
            },
            context={"phase": "LP5", "actor": "Energy Consultant"},
        ),
        UniversalRequirement(
            element_filter={"ifc_class": "IFCDOOR"},
            property_group="Pset_DoorCommon",
            property_name="IsExternal",
            constraint_def={"datatype": "IFCBOOLEAN", "cardinality": "required"},
            context={"actor": "Architect"},
        ),
        UniversalRequirement(
            element_filter={"ifc_class": "IFCBEAM"},
            property_group="Pset_BeamCommon",
            property_name="LoadBearing",
            constraint_def={"datatype": "IFCBOOLEAN", "cardinality": "required"},
            context={"phase": "LP3", "actor": "Structural Engineer"},
        ),
        UniversalRequirement(
            element_filter={"ifc_class": "IFCSPACE"},
            property_group="Pset_SpaceCommon",
            property_name="Reference",
            constraint_def={
                "datatype": "IFCLABEL",
                "cardinality": "required",
                "pattern": "^[A-Z]{2}\\d{3}$",
            },
            context={"use_case": "Space Management"},
        ),
    ]

    return export_excel(example_reqs, title="BIM Requirements Template (Example Data)")
