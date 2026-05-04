"""‚Äå‚ÅÝ‚ÄçCOBie Excel/CSV parser for BIM requirements.

Reads the 'Attribute' sheet from a COBie-formatted Excel workbook and
converts rows into UniversalRequirement objects.
"""

import logging
from pathlib import Path
from typing import Any

from app.modules.bim_requirements.parsers.base import (
    BaseRequirementParser,
    ParseResult,
    UniversalRequirement,
)

logger = logging.getLogger(__name__)

# COBie Attribute sheet columns (standard COBie 2.4 layout)
_COBIE_COLS = {
    "name": "Name",
    "created_by": "CreatedBy",
    "created_on": "CreatedOn",
    "category": "Category",
    "sheet_name": "SheetName",
    "row_name": "RowName",
    "value": "Value",
    "unit": "Unit",
    "ext_system": "ExtSystem",
    "ext_object": "ExtObject",
    "ext_identifier": "ExtIdentifier",
    "description": "Description",
    "allowed_values": "AllowedValues",
}


class COBieParser(BaseRequirementParser):
    """‚Äå‚ÅÝ‚ÄçParser for COBie Excel files (reads the Attribute sheet)."""

    FORMAT_NAME = "COBie"
    SUPPORTED_EXTENSIONS = [".xlsx", ".xls"]

    def parse(self, source: Path | str | bytes) -> ParseResult:
        """‚Äå‚ÅÝ‚ÄçParse a COBie Excel file into universal requirements."""
        result = ParseResult()
        result.metadata["format"] = self.FORMAT_NAME

        try:
            rows, headers = self._read_attribute_sheet(source)
        except Exception as exc:
            result.errors.append(
                {"row": 0, "field": "file", "msg": f"Cannot read COBie file: {exc}"}
            )
            return result

        if not rows:
            result.warnings.append(
                {"row": 0, "field": "data", "msg": "No data rows in Attribute sheet"}
            )
            return result

        # Map header names to column indices
        col_map = self._map_headers(headers)
        result.metadata["headers"] = headers
        result.metadata["row_count"] = len(rows)

        for row_idx, row in enumerate(rows, start=2):
            try:
                req = self._parse_row(row, col_map, row_idx)
                if req:
                    result.requirements.append(req)
            except Exception as exc:
                result.errors.append(
                    {"row": row_idx, "field": "", "msg": f"Error parsing row: {exc}"}
                )

        logger.info(
            "COBie parsed: %d requirements from %d rows",
            len(result.requirements),
            len(rows),
        )
        return result

    def _read_attribute_sheet(
        self, source: Path | str | bytes
    ) -> tuple[list[list[Any]], list[str]]:
        """Read the Attribute sheet from a COBie Excel file.

        Returns:
            Tuple of (data_rows, headers).
        """
        import openpyxl

        if isinstance(source, Path):
            wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
        elif isinstance(source, bytes):
            import io

            wb = openpyxl.load_workbook(
                io.BytesIO(source), read_only=True, data_only=True
            )
        else:
            raise TypeError(f"COBie parser requires a file path or bytes, got {type(source)}")

        # Find the Attribute sheet (case-insensitive)
        attr_sheet = None
        for name in wb.sheetnames:
            if name.lower() == "attribute":
                attr_sheet = wb[name]
                break

        if attr_sheet is None:
            wb.close()
            raise ValueError("No 'Attribute' sheet found in workbook")

        all_rows: list[list[Any]] = []
        for row in attr_sheet.iter_rows(values_only=True):
            all_rows.append(list(row))

        wb.close()

        if not all_rows:
            return [], []

        headers = [str(h).strip() if h else "" for h in all_rows[0]]
        data_rows = all_rows[1:]
        return data_rows, headers

    def _map_headers(self, headers: list[str]) -> dict[str, int]:
        """Map COBie column names to indices."""
        col_map: dict[str, int] = {}
        for idx, h in enumerate(headers):
            h_lower = h.lower().strip()
            for key, cobie_name in _COBIE_COLS.items():
                if h_lower == cobie_name.lower():
                    col_map[key] = idx
                    break
        return col_map

    def _get_cell(
        self, row: list[Any], col_map: dict[str, int], key: str
    ) -> str:
        """Safely get a cell value from the row."""
        idx = col_map.get(key)
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    def _parse_row(
        self,
        row: list[Any],
        col_map: dict[str, int],
        row_idx: int,
    ) -> UniversalRequirement | None:
        """Parse a single COBie Attribute row into a UniversalRequirement."""
        property_name = self._get_cell(row, col_map, "name")
        if not property_name:
            return None

        # Element filter from SheetName + RowName
        element_filter: dict[str, Any] = {}
        sheet_name = self._get_cell(row, col_map, "sheet_name")
        row_name = self._get_cell(row, col_map, "row_name")
        if sheet_name:
            element_filter["cobie_sheet"] = sheet_name
        if row_name:
            element_filter["cobie_row_name"] = row_name

        # Constraint definition
        constraint_def: dict[str, Any] = {"cardinality": "required"}

        value = self._get_cell(row, col_map, "value")
        if value:
            constraint_def["value"] = value

        unit = self._get_cell(row, col_map, "unit")
        if unit:
            constraint_def["unit"] = unit

        allowed_values = self._get_cell(row, col_map, "allowed_values")
        if allowed_values:
            # Parse semicolon or comma separated enum values
            separator = ";" if ";" in allowed_values else ","
            enums = [v.strip() for v in allowed_values.split(separator) if v.strip()]
            if enums:
                constraint_def["enum"] = enums

        # Context
        context: dict[str, Any] = {}
        category = self._get_cell(row, col_map, "category")
        if category:
            context["category"] = category

        created_by = self._get_cell(row, col_map, "created_by")
        if created_by:
            context["actor"] = created_by

        description = self._get_cell(row, col_map, "description")
        if description:
            context["description"] = description

        ext_system = self._get_cell(row, col_map, "ext_system")
        if ext_system:
            context["source"] = ext_system

        return UniversalRequirement(
            element_filter=element_filter,
            property_group=None,  # COBie does not have property groups
            property_name=property_name,
            constraint_def=constraint_def,
            context=context if context else None,
        )
