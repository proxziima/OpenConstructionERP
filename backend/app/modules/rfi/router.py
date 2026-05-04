"""‚Äå‚ÅÝ‚ÄçRFI API routes.

Endpoints:
    GET    /                    - List RFIs for a project
    POST   /                    - Create RFI
    GET    /export              - Export RFI log as Excel
    GET    /{rfi_id}            - Get single RFI
    PATCH  /{rfi_id}            - Update RFI
    DELETE /{rfi_id}            - Delete RFI
    POST   /{rfi_id}/respond    - Record official response
    POST   /{rfi_id}/close      - Close RFI
"""

import io
import logging
import uuid
from datetime import UTC, datetime

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse

from app.core.bulk_ops import BulkDeleteRequest, BulkStatusRequest
from app.dependencies import CurrentUserId, RequirePermission, SessionDep, verify_project_access
from app.modules.rfi.schemas import (
    RFICreate,
    RFIRespondRequest,
    RFIResponse,
    RFIStatsResponse,
    RFIUpdate,
)
from app.modules.rfi.service import RFIService

router = APIRouter()
logger = logging.getLogger(__name__)


def _get_service(session: SessionDep) -> RFIService:
    return RFIService(session)


def _compute_rfi_fields(item: object) -> tuple[bool, int]:
    """‚Äå‚ÅÝ‚ÄçCompute is_overdue and days_open for an RFI item."""
    now = datetime.now(UTC)

    # days_open: from created_at to now (or responded_at if answered/closed)
    days_open = 0
    created_at = getattr(item, "created_at", None)
    if created_at is not None:
        end = now
        status = getattr(item, "status", "")
        responded_at = getattr(item, "responded_at", None)
        if status in ("answered", "closed") and responded_at:
            try:
                end = datetime.fromisoformat(str(responded_at))
                if end.tzinfo is None:
                    end = end.replace(tzinfo=UTC)
            except (ValueError, TypeError):
                end = now
        start = created_at
        if start.tzinfo is None:
            start = start.replace(tzinfo=UTC)
        if end.tzinfo is None:
            end = end.replace(tzinfo=UTC)
        try:
            days_open = max(0, (end - start).days)
        except TypeError:
            days_open = 0

    # is_overdue: open/draft status + response_due_date in the past
    is_overdue = False
    status = getattr(item, "status", "")
    response_due_date = getattr(item, "response_due_date", None)
    if status in ("draft", "open") and response_due_date:
        try:
            due = datetime.fromisoformat(str(response_due_date))
            if due.tzinfo is None:
                due = due.replace(tzinfo=UTC)
            is_overdue = now > due
        except (ValueError, TypeError):
            is_overdue = False

    return is_overdue, days_open


def _to_response(item: object) -> RFIResponse:
    is_overdue, days_open = _compute_rfi_fields(item)
    return RFIResponse(
        id=item.id,  # type: ignore[attr-defined]
        project_id=item.project_id,  # type: ignore[attr-defined]
        rfi_number=item.rfi_number,  # type: ignore[attr-defined]
        subject=item.subject,  # type: ignore[attr-defined]
        question=item.question,  # type: ignore[attr-defined]
        raised_by=item.raised_by,  # type: ignore[attr-defined]
        assigned_to=str(item.assigned_to) if item.assigned_to else None,  # type: ignore[attr-defined]
        status=item.status,  # type: ignore[attr-defined]
        ball_in_court=str(item.ball_in_court) if item.ball_in_court else None,  # type: ignore[attr-defined]
        official_response=item.official_response,  # type: ignore[attr-defined]
        responded_by=str(item.responded_by) if item.responded_by else None,  # type: ignore[attr-defined]
        responded_at=item.responded_at,  # type: ignore[attr-defined]
        cost_impact=item.cost_impact,  # type: ignore[attr-defined]
        cost_impact_value=item.cost_impact_value,  # type: ignore[attr-defined]
        schedule_impact=item.schedule_impact,  # type: ignore[attr-defined]
        schedule_impact_days=item.schedule_impact_days,  # type: ignore[attr-defined]
        date_required=item.date_required,  # type: ignore[attr-defined]
        response_due_date=item.response_due_date,  # type: ignore[attr-defined]
        linked_drawing_ids=item.linked_drawing_ids or [],  # type: ignore[attr-defined]
        change_order_id=item.change_order_id,  # type: ignore[attr-defined]
        created_by=item.created_by,  # type: ignore[attr-defined]
        metadata=getattr(item, "metadata_", {}),
        created_at=item.created_at,  # type: ignore[attr-defined]
        updated_at=item.updated_at,  # type: ignore[attr-defined]
        is_overdue=is_overdue,
        days_open=days_open,
    )


@router.get(
    "/",
    response_model=list[RFIResponse],
    dependencies=[Depends(RequirePermission("rfi.read"))],
)
async def list_rfis(
    user_id: CurrentUserId,
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    status_filter: str | None = Query(default=None, alias="status"),
    search: str | None = Query(
        default=None,
        max_length=200,
        description="Free-text search across subject, question, response, and RFI number.",
    ),
    service: RFIService = Depends(_get_service),
) -> list[RFIResponse]:
    """‚Äå‚ÅÝ‚ÄçList RFIs for a project."""
    await verify_project_access(project_id, user_id, session)
    rfis, _ = await service.list_rfis(
        project_id,
        offset=offset,
        limit=limit,
        status_filter=status_filter,
        search=search,
    )
    return [_to_response(r) for r in rfis]


@router.post("/", response_model=RFIResponse, status_code=201)
async def create_rfi(
    data: RFICreate,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("rfi.create")),
    service: RFIService = Depends(_get_service),
) -> RFIResponse:
    """Create a new RFI."""
    await verify_project_access(data.project_id, user_id, session)
    rfi = await service.create_rfi(data, user_id=user_id)
    return _to_response(rfi)


@router.get(
    "/stats/",
    response_model=RFIStatsResponse,
    dependencies=[Depends(RequirePermission("rfi.read"))],
)
async def rfi_stats(
    user_id: CurrentUserId,
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
    service: RFIService = Depends(_get_service),
) -> RFIStatsResponse:
    """Return summary statistics for RFIs in a project.

    Computes total, open, overdue, avg response time, and breakdown by status.
    """
    await verify_project_access(project_id, user_id, session)
    return await service.get_stats(project_id)


@router.get(
    "/export/",
    dependencies=[Depends(RequirePermission("rfi.read"))],
)
async def export_rfi_log(
    _user: CurrentUserId,
    session: SessionDep,
    project_id: uuid.UUID = Query(...),
) -> StreamingResponse:
    """Export RFI log for a project as Excel."""
    await verify_project_access(project_id, _user, session)
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from sqlalchemy import select

    from app.modules.rfi.models import RFI

    result = await session.execute(
        select(RFI)
        .where(RFI.project_id == project_id)
        .order_by(RFI.rfi_number)
        .limit(50000)
    )
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "RFI Log"

    headers = [
        "RFI #",
        "Subject",
        "Status",
        "Raised By",
        "Assigned To",
        "Ball-in-Court",
        "Date Required",
        "Response Due",
        "Days Open",
        "Cost Impact",
        "Schedule Impact",
        "Response",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    now = datetime.now(UTC)
    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.rfi_number)
        ws.cell(row=row_idx, column=2, value=item.subject)
        ws.cell(row=row_idx, column=3, value=item.status)
        ws.cell(row=row_idx, column=4, value=str(item.raised_by) if item.raised_by else "")
        ws.cell(row=row_idx, column=5, value=str(item.assigned_to) if item.assigned_to else "")
        ws.cell(row=row_idx, column=6, value=str(item.ball_in_court) if item.ball_in_court else "")
        ws.cell(row=row_idx, column=7, value=item.date_required or "")
        ws.cell(row=row_idx, column=8, value=item.response_due_date or "")
        # Days open: from created_at to now (or responded_at if closed)
        days_open = 0
        if item.created_at:
            start = item.created_at if hasattr(item.created_at, "date") else now
            end = now
            if item.responded_at:
                try:
                    end = datetime.fromisoformat(str(item.responded_at))
                except (ValueError, TypeError):
                    end = now
            try:
                days_open = max(0, (end - start).days)
            except TypeError:
                days_open = 0
        ws.cell(row=row_idx, column=9, value=days_open)
        ws.cell(
            row=row_idx,
            column=10,
            value=f"Yes ({item.cost_impact_value})" if item.cost_impact else "No",
        )
        ws.cell(
            row=row_idx,
            column=11,
            value=f"Yes ({item.schedule_impact_days}d)" if item.schedule_impact else "No",
        )
        ws.cell(row=row_idx, column=12, value=item.official_response or "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="rfi_log.xlsx"'},
    )


# ‚îÄ‚îÄ Bulk operations (must be BEFORE parametric /{rfi_id}) ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ‚îÄ


@router.post(
    "/batch/delete/",
    status_code=200,
    dependencies=[Depends(RequirePermission("rfi.delete"))],
)
async def batch_delete_rfis(
    body: BulkDeleteRequest,
    user_id: CurrentUserId,
    session: SessionDep,
) -> dict:
    """Delete multiple RFIs in one request."""
    from sqlalchemy import select as _select

    from app.core.bulk_ops import bulk_delete
    from app.modules.projects.repository import ProjectRepository
    from app.modules.rfi.models import RFI

    proj_repo = ProjectRepository(session)
    owned_projects, _ = await proj_repo.list_for_user(
        owner_id=user_id, offset=0, limit=10000, exclude_archived=False
    )
    owned_project_ids = {str(p.id) for p in owned_projects}

    rows = (await session.execute(
        _select(RFI.id, RFI.project_id).where(RFI.id.in_(body.ids))
    )).all()
    allowed = [r[0] for r in rows if str(r[1]) in owned_project_ids]

    deleted = await bulk_delete(session, RFI, allowed)
    logger.info(
        "Bulk delete RFIs: requested=%d deleted=%d user=%s",
        len(body.ids), deleted, user_id,
    )
    return {"requested": len(body.ids), "deleted": deleted}


@router.patch(
    "/batch/status/",
    status_code=200,
    dependencies=[Depends(RequirePermission("rfi.update"))],
)
async def batch_update_rfi_status(
    body: BulkStatusRequest,
    user_id: CurrentUserId,
    session: SessionDep,
) -> dict:
    """Bulk-update status on multiple RFIs."""
    from sqlalchemy import select as _select

    from app.core.bulk_ops import bulk_update_status
    from app.modules.projects.repository import ProjectRepository
    from app.modules.rfi.models import RFI

    allowed_statuses = {"draft", "open", "answered", "closed", "void"}
    if body.status not in allowed_statuses:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid status. Allowed: {sorted(allowed_statuses)}",
        )

    proj_repo = ProjectRepository(session)
    owned_projects, _ = await proj_repo.list_for_user(
        owner_id=user_id, offset=0, limit=10000, exclude_archived=False
    )
    owned_project_ids = {str(p.id) for p in owned_projects}

    rows = (await session.execute(
        _select(RFI.id, RFI.project_id).where(RFI.id.in_(body.ids))
    )).all()
    allowed_ids = [r[0] for r in rows if str(r[1]) in owned_project_ids]

    updated = await bulk_update_status(
        session, RFI, allowed_ids, body.status, allowed_statuses=allowed_statuses
    )
    logger.info(
        "Bulk update RFI status: requested=%d updated=%d user=%s",
        len(body.ids), updated, user_id,
    )
    return {"requested": len(body.ids), "updated": updated, "status": body.status}


@router.get(
    "/{rfi_id}",
    response_model=RFIResponse,
    dependencies=[Depends(RequirePermission("rfi.read"))],
)
async def get_rfi(
    rfi_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId,
    service: RFIService = Depends(_get_service),
) -> RFIResponse:
    """Get a single RFI."""
    rfi = await service.get_rfi(rfi_id)
    await verify_project_access(rfi.project_id, str(user_id), session)
    return _to_response(rfi)


@router.patch("/{rfi_id}", response_model=RFIResponse)
async def update_rfi(
    rfi_id: uuid.UUID,
    data: RFIUpdate,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("rfi.update")),
    service: RFIService = Depends(_get_service),
) -> RFIResponse:
    """Update an RFI."""
    existing = await service.get_rfi(rfi_id)
    await verify_project_access(existing.project_id, str(user_id), session)
    rfi = await service.update_rfi(rfi_id, data)
    return _to_response(rfi)


@router.delete("/{rfi_id}", status_code=204)
async def delete_rfi(
    rfi_id: uuid.UUID,
    session: SessionDep,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("rfi.delete")),
    service: RFIService = Depends(_get_service),
) -> None:
    """Delete an RFI."""
    existing = await service.get_rfi(rfi_id)
    await verify_project_access(existing.project_id, str(user_id), session)
    await service.delete_rfi(rfi_id)


@router.post("/{rfi_id}/respond/", response_model=RFIResponse)
async def respond_to_rfi(
    rfi_id: uuid.UUID,
    body: RFIRespondRequest,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("rfi.update")),
    service: RFIService = Depends(_get_service),
) -> RFIResponse:
    """Record an official response to an RFI."""
    rfi = await service.respond_to_rfi(rfi_id, body.official_response, responded_by=user_id)
    return _to_response(rfi)


@router.post("/{rfi_id}/create-variation/", status_code=201)
async def create_variation_from_rfi(
    rfi_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    _perm: None = Depends(RequirePermission("rfi.update")),
    service: RFIService = Depends(_get_service),
) -> dict:
    """Create a change order/variation pre-filled from an RFI with cost impact.

    The RFI must have cost_impact=True and be in 'answered' or 'closed' status.
    Pre-fills the change order with the RFI's subject, question, and cost impact value.
    """
    rfi = await service.get_rfi(rfi_id)

    if not rfi.cost_impact:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="RFI has no cost impact ‚Äî cannot create a variation.",
        )

    if rfi.status not in ("answered", "closed"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="RFI must be answered or closed before creating a variation.",
        )

    # Lazy import changeorders module
    try:
        from app.modules.changeorders.models import ChangeOrder
        from app.modules.changeorders.repository import ChangeOrderRepository

        repo = ChangeOrderRepository(session)
        count = await repo.count_for_project(rfi.project_id)
        code = f"CO-{count + 1:03d}"

        description_parts = [
            f"Variation from RFI {rfi.rfi_number}: {rfi.subject}",
            "",
            "Question:",
            rfi.question,
        ]
        if rfi.official_response:
            description_parts.extend(["", "Response:", rfi.official_response])

        order = ChangeOrder(
            project_id=rfi.project_id,
            code=code,
            title=f"Variation: {rfi.subject}",
            description="\n".join(description_parts),
            reason_category="client_request",
            cost_impact=rfi.cost_impact_value or "0",
            schedule_impact_days=rfi.schedule_impact_days or 0,
            metadata_={
                "source": "rfi",
                "rfi_id": str(rfi_id),
                "rfi_number": rfi.rfi_number,
            },
        )
        session.add(order)
        await session.flush()

        # Link the change order back to the RFI
        rfi.change_order_id = str(order.id)
        await session.flush()

        logger.info(
            "Created change order %s from RFI %s",
            code,
            rfi_id,
        )
        return {
            "change_order_id": str(order.id),
            "code": code,
            "rfi_id": str(rfi_id),
            "title": order.title,
        }
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Change orders module is not available.",
        )
    except Exception as exc:
        logger.exception("Failed to create variation from RFI %s: %s", rfi_id, exc)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to create change order from RFI.",
        )


@router.post(
    "/{rfi_id}/close/",
    response_model=RFIResponse,
    dependencies=[Depends(RequirePermission("rfi.update"))],
)
async def close_rfi(
    rfi_id: uuid.UUID,
    user_id: CurrentUserId,
    service: RFIService = Depends(_get_service),
) -> RFIResponse:
    """Close an RFI."""
    rfi = await service.close_rfi(rfi_id, closed_by=user_id)
    return _to_response(rfi)
