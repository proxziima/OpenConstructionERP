"""Tasks API routes.

Endpoints:
    GET    /                    - List tasks for a project
    POST   /                    - Create task
    GET    /my-tasks             - List tasks for the current user
    GET    /export               - Export tasks as Excel file
    GET    /{task_id}            - Get single task
    PATCH  /{task_id}            - Update task
    DELETE /{task_id}            - Delete task
    POST   /{task_id}/complete   - Mark task as completed
"""

import io
import logging
import uuid

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

from app.dependencies import CurrentUserId, RequirePermission, SessionDep
from app.modules.tasks.schemas import (
    TaskCompleteRequest,
    TaskCreate,
    TaskResponse,
    TaskUpdate,
)
from app.modules.tasks.service import TaskService

router = APIRouter()
logger = logging.getLogger(__name__)


def _get_service(session: SessionDep) -> TaskService:
    return TaskService(session)


def _to_response(item: object) -> TaskResponse:
    return TaskResponse(
        id=item.id,  # type: ignore[attr-defined]
        project_id=item.project_id,  # type: ignore[attr-defined]
        task_type=item.task_type,  # type: ignore[attr-defined]
        title=item.title,  # type: ignore[attr-defined]
        description=item.description,  # type: ignore[attr-defined]
        checklist=item.checklist or [],  # type: ignore[attr-defined]
        responsible_id=str(item.responsible_id) if item.responsible_id else None,  # type: ignore[attr-defined]
        persons_involved=item.persons_involved or [],  # type: ignore[attr-defined]
        due_date=item.due_date,  # type: ignore[attr-defined]
        milestone_id=item.milestone_id,  # type: ignore[attr-defined]
        meeting_id=item.meeting_id,  # type: ignore[attr-defined]
        status=item.status,  # type: ignore[attr-defined]
        priority=item.priority,  # type: ignore[attr-defined]
        result=item.result,  # type: ignore[attr-defined]
        is_private=item.is_private,  # type: ignore[attr-defined]
        created_by=item.created_by,  # type: ignore[attr-defined]
        metadata=getattr(item, "metadata_", {}),
        created_at=item.created_at,  # type: ignore[attr-defined]
        updated_at=item.updated_at,  # type: ignore[attr-defined]
    )


@router.get("/", response_model=list[TaskResponse])
async def list_tasks(
    project_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    type_filter: str | None = Query(default=None, alias="type"),
    status_filter: str | None = Query(default=None, alias="status"),
    priority: str | None = Query(default=None),
    responsible_id: str | None = Query(default=None),
    service: TaskService = Depends(_get_service),
) -> list[TaskResponse]:
    """List tasks for a project with optional filters."""
    tasks, _ = await service.list_tasks(
        project_id,
        offset=offset,
        limit=limit,
        task_type=type_filter,
        status_filter=status_filter,
        priority=priority,
        responsible_id=responsible_id,
    )
    return [_to_response(t) for t in tasks]


@router.get("/my-tasks", response_model=list[TaskResponse])
async def my_tasks(
    user_id: CurrentUserId,
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    status_filter: str | None = Query(default=None, alias="status"),
    service: TaskService = Depends(_get_service),
) -> list[TaskResponse]:
    """List tasks assigned to the current user across all projects."""
    tasks, _ = await service.list_my_tasks(
        user_id,
        offset=offset,
        limit=limit,
        status_filter=status_filter,
    )
    return [_to_response(t) for t in tasks]


@router.post("/", response_model=TaskResponse, status_code=201)
async def create_task(
    data: TaskCreate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("tasks.create")),
    service: TaskService = Depends(_get_service),
) -> TaskResponse:
    """Create a new task."""
    task = await service.create_task(data, user_id=user_id)
    return _to_response(task)


# ── Export tasks as Excel ────────────────────────────────────────────────────


@router.get("/export")
async def export_tasks(
    project_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TaskService = Depends(_get_service),
) -> StreamingResponse:
    """Export tasks for a project as Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    tasks, _ = await service.list_tasks(project_id, offset=0, limit=10000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = [
        "Title",
        "Type",
        "Status",
        "Priority",
        "Assignee",
        "Due Date",
        "Created",
        "Checklist Progress",
    ]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    for row_idx, task in enumerate(tasks, 2):
        ws.cell(row=row_idx, column=1, value=task.title)  # type: ignore[attr-defined]
        ws.cell(row=row_idx, column=2, value=task.task_type)  # type: ignore[attr-defined]
        ws.cell(row=row_idx, column=3, value=task.status)  # type: ignore[attr-defined]
        ws.cell(row=row_idx, column=4, value=task.priority)  # type: ignore[attr-defined]
        ws.cell(
            row=row_idx,
            column=5,
            value=str(task.responsible_id) if task.responsible_id else "",  # type: ignore[attr-defined]
        )
        ws.cell(row=row_idx, column=6, value=task.due_date)  # type: ignore[attr-defined]
        ws.cell(
            row=row_idx,
            column=7,
            value=str(task.created_at) if task.created_at else "",  # type: ignore[attr-defined]
        )
        # Checklist progress
        checklist = task.checklist or []  # type: ignore[attr-defined]
        total = len(checklist)
        done = sum(1 for c in checklist if isinstance(c, dict) and c.get("completed"))
        ws.cell(
            row=row_idx,
            column=8,
            value=f"{done}/{total}" if total > 0 else "",
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="tasks_export.xlsx"'},
    )


@router.get("/{task_id}", response_model=TaskResponse)
async def get_task(
    task_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: TaskService = Depends(_get_service),
) -> TaskResponse:
    """Get a single task."""
    task = await service.get_task(task_id)
    return _to_response(task)


@router.patch("/{task_id}", response_model=TaskResponse)
async def update_task(
    task_id: uuid.UUID,
    data: TaskUpdate,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("tasks.update")),
    service: TaskService = Depends(_get_service),
) -> TaskResponse:
    """Update a task."""
    task = await service.update_task(task_id, data)
    return _to_response(task)


@router.delete("/{task_id}", status_code=204)
async def delete_task(
    task_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("tasks.delete")),
    service: TaskService = Depends(_get_service),
) -> None:
    """Delete a task."""
    await service.delete_task(task_id)


@router.post("/{task_id}/complete", response_model=TaskResponse)
async def complete_task(
    task_id: uuid.UUID,
    body: TaskCompleteRequest | None = None,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("tasks.update")),
    service: TaskService = Depends(_get_service),
) -> TaskResponse:
    """Mark a task as completed with optional result text."""
    result = body.result if body else None
    task = await service.complete_task(task_id, result=result)
    return _to_response(task)
