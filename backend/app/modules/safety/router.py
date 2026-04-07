"""Safety API routes.

Endpoints:
    GET    /incidents                   - List incidents for a project
    POST   /incidents                   - Create incident
    GET    /incidents/{id}              - Get single incident
    PATCH  /incidents/{id}              - Update incident
    DELETE /incidents/{id}              - Delete incident
    GET    /incidents/export            - Export incidents as Excel
    GET    /observations                - List observations for a project
    POST   /observations                - Create observation
    GET    /observations/{id}           - Get single observation
    PATCH  /observations/{id}           - Update observation
    DELETE /observations/{id}           - Delete observation
    GET    /observations/export         - Export observations as Excel
"""

import io
import logging
import uuid

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

from app.dependencies import CurrentUserId, RequirePermission, SessionDep
from app.modules.safety.schemas import (
    IncidentCreate,
    IncidentResponse,
    IncidentUpdate,
    ObservationCreate,
    ObservationResponse,
    ObservationUpdate,
)
from app.modules.safety.service import SafetyService

router = APIRouter()
logger = logging.getLogger(__name__)


def _get_service(session: SessionDep) -> SafetyService:
    return SafetyService(session)


def _incident_to_response(item: object) -> IncidentResponse:
    return IncidentResponse(
        id=item.id,  # type: ignore[attr-defined]
        project_id=item.project_id,  # type: ignore[attr-defined]
        incident_number=item.incident_number,  # type: ignore[attr-defined]
        incident_date=item.incident_date,  # type: ignore[attr-defined]
        location=item.location,  # type: ignore[attr-defined]
        incident_type=item.incident_type,  # type: ignore[attr-defined]
        description=item.description,  # type: ignore[attr-defined]
        injured_person_details=item.injured_person_details,  # type: ignore[attr-defined]
        treatment_type=item.treatment_type,  # type: ignore[attr-defined]
        days_lost=item.days_lost,  # type: ignore[attr-defined]
        root_cause=item.root_cause,  # type: ignore[attr-defined]
        corrective_actions=item.corrective_actions or [],  # type: ignore[attr-defined]
        reported_to_regulator=item.reported_to_regulator,  # type: ignore[attr-defined]
        status=item.status,  # type: ignore[attr-defined]
        created_by=item.created_by,  # type: ignore[attr-defined]
        metadata=getattr(item, "metadata_", {}),
        created_at=item.created_at,  # type: ignore[attr-defined]
        updated_at=item.updated_at,  # type: ignore[attr-defined]
    )


def _observation_to_response(item: object) -> ObservationResponse:
    return ObservationResponse(
        id=item.id,  # type: ignore[attr-defined]
        project_id=item.project_id,  # type: ignore[attr-defined]
        observation_number=item.observation_number,  # type: ignore[attr-defined]
        observation_type=item.observation_type,  # type: ignore[attr-defined]
        description=item.description,  # type: ignore[attr-defined]
        location=item.location,  # type: ignore[attr-defined]
        severity=item.severity,  # type: ignore[attr-defined]
        likelihood=item.likelihood,  # type: ignore[attr-defined]
        risk_score=item.risk_score,  # type: ignore[attr-defined]
        immediate_action=item.immediate_action,  # type: ignore[attr-defined]
        corrective_action=item.corrective_action,  # type: ignore[attr-defined]
        status=item.status,  # type: ignore[attr-defined]
        created_by=item.created_by,  # type: ignore[attr-defined]
        metadata=getattr(item, "metadata_", {}),
        created_at=item.created_at,  # type: ignore[attr-defined]
        updated_at=item.updated_at,  # type: ignore[attr-defined]
    )


# ── Incidents ────────────────────────────────────────────────────────────


@router.get("/incidents", response_model=list[IncidentResponse])
async def list_incidents(
    project_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    type_filter: str | None = Query(default=None, alias="type"),
    status_filter: str | None = Query(default=None, alias="status"),
    service: SafetyService = Depends(_get_service),
) -> list[IncidentResponse]:
    items, _ = await service.list_incidents(
        project_id,
        offset=offset,
        limit=limit,
        incident_type=type_filter,
        status_filter=status_filter,
    )
    return [_incident_to_response(i) for i in items]


@router.post("/incidents", response_model=IncidentResponse, status_code=201)
async def create_incident(
    data: IncidentCreate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("safety.create")),
    service: SafetyService = Depends(_get_service),
) -> IncidentResponse:
    incident = await service.create_incident(data, user_id=user_id)
    return _incident_to_response(incident)


@router.get("/incidents/export")
async def export_incidents(
    project_id: uuid.UUID = Query(...),
    session: SessionDep = None,  # type: ignore[assignment]
    _user: CurrentUserId = None,  # type: ignore[assignment]
) -> StreamingResponse:
    """Export safety incidents for a project as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from sqlalchemy import select

    from app.modules.safety.models import SafetyIncident

    result = await session.execute(
        select(SafetyIncident)
        .where(SafetyIncident.project_id == project_id)
        .order_by(SafetyIncident.incident_number)
        .limit(50000)
    )
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Safety Incidents"

    headers = [
        "Incident #",
        "Date",
        "Type",
        "Location",
        "Description",
        "Severity",
        "Treatment",
        "Days Lost",
        "Root Cause",
        "Status",
        "Reported to Regulator",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.incident_number)
        ws.cell(row=row_idx, column=2, value=item.incident_date)
        ws.cell(row=row_idx, column=3, value=item.incident_type)
        ws.cell(row=row_idx, column=4, value=item.location or "")
        ws.cell(row=row_idx, column=5, value=item.description)
        # Severity is stored in metadata or injured_person_details; use type as proxy
        severity = ""
        if isinstance(item.injured_person_details, dict):
            severity = item.injured_person_details.get("severity", "")
        ws.cell(row=row_idx, column=6, value=severity)
        ws.cell(row=row_idx, column=7, value=item.treatment_type or "")
        ws.cell(row=row_idx, column=8, value=item.days_lost)
        ws.cell(row=row_idx, column=9, value=item.root_cause or "")
        ws.cell(row=row_idx, column=10, value=item.status)
        ws.cell(row=row_idx, column=11, value="Yes" if item.reported_to_regulator else "No")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="safety_incidents.xlsx"'},
    )


@router.get("/incidents/{incident_id}", response_model=IncidentResponse)
async def get_incident(
    incident_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: SafetyService = Depends(_get_service),
) -> IncidentResponse:
    incident = await service.get_incident(incident_id)
    return _incident_to_response(incident)


@router.patch("/incidents/{incident_id}", response_model=IncidentResponse)
async def update_incident(
    incident_id: uuid.UUID,
    data: IncidentUpdate,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("safety.update")),
    service: SafetyService = Depends(_get_service),
) -> IncidentResponse:
    incident = await service.update_incident(incident_id, data)
    return _incident_to_response(incident)


@router.delete("/incidents/{incident_id}", status_code=204)
async def delete_incident(
    incident_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("safety.delete")),
    service: SafetyService = Depends(_get_service),
) -> None:
    await service.delete_incident(incident_id)


# ── Observations ─────────────────────────────────────────────────────────


@router.get("/observations", response_model=list[ObservationResponse])
async def list_observations(
    project_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    type_filter: str | None = Query(default=None, alias="type"),
    status_filter: str | None = Query(default=None, alias="status"),
    service: SafetyService = Depends(_get_service),
) -> list[ObservationResponse]:
    items, _ = await service.list_observations(
        project_id,
        offset=offset,
        limit=limit,
        observation_type=type_filter,
        status_filter=status_filter,
    )
    return [_observation_to_response(i) for i in items]


@router.post("/observations", response_model=ObservationResponse, status_code=201)
async def create_observation(
    data: ObservationCreate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("safety.create")),
    service: SafetyService = Depends(_get_service),
) -> ObservationResponse:
    observation = await service.create_observation(data, user_id=user_id)
    return _observation_to_response(observation)


@router.get("/observations/export")
async def export_observations(
    project_id: uuid.UUID = Query(...),
    session: SessionDep = None,  # type: ignore[assignment]
    _user: CurrentUserId = None,  # type: ignore[assignment]
) -> StreamingResponse:
    """Export safety observations for a project as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from sqlalchemy import select

    from app.modules.safety.models import SafetyObservation

    result = await session.execute(
        select(SafetyObservation)
        .where(SafetyObservation.project_id == project_id)
        .order_by(SafetyObservation.observation_number)
        .limit(50000)
    )
    items = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Safety Observations"

    headers = [
        "Observation #",
        "Date",
        "Type",
        "Location",
        "Description",
        "Severity",
        "Likelihood",
        "Risk Score",
        "Risk Tier",
        "Status",
        "Corrective Action",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.observation_number)
        ws.cell(row=row_idx, column=2, value=str(item.created_at) if item.created_at else "")
        ws.cell(row=row_idx, column=3, value=item.observation_type)
        ws.cell(row=row_idx, column=4, value=item.location or "")
        ws.cell(row=row_idx, column=5, value=item.description)
        ws.cell(row=row_idx, column=6, value=item.severity)
        ws.cell(row=row_idx, column=7, value=item.likelihood)
        ws.cell(row=row_idx, column=8, value=item.risk_score)
        # Risk tier derived from risk score
        risk_tier = "Low"
        if item.risk_score > 15:
            risk_tier = "Critical"
        elif item.risk_score > 10:
            risk_tier = "High"
        elif item.risk_score > 5:
            risk_tier = "Medium"
        ws.cell(row=row_idx, column=9, value=risk_tier)
        ws.cell(row=row_idx, column=10, value=item.status)
        ws.cell(row=row_idx, column=11, value=item.corrective_action or "")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="safety_observations.xlsx"'},
    )


@router.get("/observations/{observation_id}", response_model=ObservationResponse)
async def get_observation(
    observation_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    service: SafetyService = Depends(_get_service),
) -> ObservationResponse:
    observation = await service.get_observation(observation_id)
    return _observation_to_response(observation)


@router.patch("/observations/{observation_id}", response_model=ObservationResponse)
async def update_observation(
    observation_id: uuid.UUID,
    data: ObservationUpdate,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("safety.update")),
    service: SafetyService = Depends(_get_service),
) -> ObservationResponse:
    observation = await service.update_observation(observation_id, data)
    return _observation_to_response(observation)


@router.delete("/observations/{observation_id}", status_code=204)
async def delete_observation(
    observation_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("safety.delete")),
    service: SafetyService = Depends(_get_service),
) -> None:
    await service.delete_observation(observation_id)
